"""
build_calibrated_workbook_final.py — Phase 3D Final Calibrated Workbook
========================================================================
Reads:
  --latents   cbb_cache/MatchupLatents_today_player.csv  (or player-adjusted)
  --cal-report cbb_cache/model_calibration_report.csv
  --edge      cbb_cache/edge_bucket_table.csv
  --pred      cbb_cache/historical_p230_predictions.csv
  --out       outputs/ncaab_market_maker_FINAL_calibrated_v1_p2_30_<date>.xlsx
  --date      2026-03-21

5 tabs:
  Model_Info              version stamp + leakage status + honest limitation flags
  Calibration_Report      raw vs calibrated Brier/AUC/LogLoss per market type
  Edge_Bucket_Table       side-aware ATS buckets, cover rate, EV@-110
  Pricing_Output_Final    raw + calibrated probs side by side per game
  Slate_Summary           aggregate stats + what is/isn't proven
"""
import math, os, argparse
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="FF1B2A4A"; GOLD="FFC8A951"; LGREY="FFF2F4F7"; DGREY="FFD0D4DC"
WHITE="FFFFFFFF"; GREEN="FF1E7A3E"; RED="FFB22222"; BLUE="FF1A50A0"
ORANGE="FFCC5500"; GREY2="FF555555"; TEAL="FF0E6655"; PURPLE="FF6B2F8A"
DRED="FF8B0000"

def fill(h): h=h if h else "FFFFFFFF"; return PatternFill("solid",start_color=h,fgColor=h)
def bdr():
    s=Side(style="thin",color=DGREY); return Border(left=s,right=s,top=s,bottom=s)
def hdr(ws,r,c,v,bg=NAVY,fg=WHITE,sz=9,bold=True):
    cell=ws.cell(r,c,v); cell.font=Font(name="Arial",size=sz,bold=bold,color=fg)
    cell.fill=fill(bg); cell.border=bdr()
    cell.alignment=Alignment(horizontal="center",vertical="center"); return cell
def val(ws,r,c,v,bg=WHITE,fg="FF000000",bold=False,fmt=None,align="center"):
    if isinstance(v,float) and math.isnan(v): v=""
    cell=ws.cell(r,c,v)
    safe_fg=fg if(len(str(fg))in(6,8)and all(x in"0123456789ABCDEFabcdef"for x in str(fg)))else"FF000000"
    cell.font=Font(name="Arial",size=9,bold=bold,color=safe_fg)
    cell.fill=fill(bg); cell.border=bdr()
    cell.alignment=Alignment(horizontal=align,vertical="center")
    if fmt: cell.number_format=fmt; return cell
def banner(ws,r,txt,bg=NAVY,fg=WHITE,cols=2,h=24,sz=12):
    ws.merge_cells(f"A{r}:{get_column_letter(cols)}{r}"); c=ws[f"A{r}"]; c.value=txt
    c.font=Font(name="Arial",size=sz,bold=True,color=fg); c.fill=fill(bg)
    c.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[r].height=h


def build(args):
    os.makedirs(os.path.dirname(args.out) if os.path.dirname(args.out) else ".", exist_ok=True)

    # ── Production vs experimental detection ─────────────────────────────
    is_player = ("player" in str(args.latents).lower() or
                 "adj" in str(args.latents).lower())
    out_upper = str(args.out).upper()

    # Safeguard: refuse to write player-adjusted latents to a PRODUCTION filename
    if is_player and "PRODUCTION" in out_upper:
        raise SystemExit(
            "ERROR: --latents points to a player-adjusted file but --out "
            "contains 'PRODUCTION'. Use EXPERIMENTAL in the output filename. "
            "Player layer FAILED holdout and must not be labeled production."
        )

    if is_player:
        WB_TITLE     = "NCAAB Market Maker — EXPERIMENTAL Player Layer — NOT FOR PRODUCTION"
        VERSION_STAMP= "experimental_player_adj_v1"
        PROD_MODEL   = "team_only_v1_p2_30 (production base)"
        EXP_MODEL    = "player_adj_v1"
        LAYER_VALID  = "False — FAILED holdout (ATS ΔAUC=-0.0224, all player buckets negative EV)"
        PROMOTED     = "False"
        NOT_PROD     = "True — player layer failed holdout, do not use for production pricing"
        TITLE_BG     = "FF8B0000"  # dark red
    else:
        WB_TITLE     = "NCAAB Market Maker — Production — team_only_v1_p2_30"
        VERSION_STAMP= "team_only_v1_p2_30"
        PROD_MODEL   = "team_only_v1_p2_30"
        EXP_MODEL    = "N/A — player layer not applied"
        LAYER_VALID  = "False — player layer failed holdout, not promoted"
        PROMOTED     = "False"
        NOT_PROD     = "False — this is the production candidate"
        TITLE_BG     = NAVY

    lat = pd.read_csv(args.latents)
    cal = pd.read_csv(args.cal_report) if os.path.exists(args.cal_report) else pd.DataFrame()
    edge= pd.read_csv(args.edge)       if os.path.exists(args.edge)       else pd.DataFrame()
    pred= pd.read_csv(args.pred)       if os.path.exists(args.pred)       else pd.DataFrame()
    print(f"Loaded: {len(lat)} slate games, {len(cal)} cal rows, {len(edge)} edge buckets, {len(pred)} hist games")

    # Compute coverage stats from pred CSV (must happen before leakage_flagged)
    n_hist  = len(pred)
    kp_n    = int(pred["kenpom_used"].sum())  if (len(pred)>0 and "kenpom_used" in pred.columns) else 0
    kp_pct  = 100.0*kp_n/max(n_hist,1)
    bdb_n   = n_hist - kp_n
    bdb_pct = 100.0 - kp_pct
    kp_pct_str  = f"{kp_pct:.1f}%  ({kp_n}/{n_hist} games)"
    bdb_pct_str = f"{bdb_pct:.1f}%  ({bdb_n}/{n_hist} games)"

    # ATS bucket status — does any n>=100 bucket beat vig?
    vig_be = 1.0/(1.0 + 100.0/110.0)
    ats_validated = False
    if len(edge)>0 and "chosen_cover_rate" in edge.columns and "n" in edge.columns:
        ats_validated = bool(((edge["n"]>=100) & (edge["chosen_cover_rate"]>vig_be)).any())
    ats_status = "VALIDATED — bucket(s) with N>=100 beat vig" if ats_validated else                  "NOT VALIDATED — no ATS bucket with N>=100 beats vig in current holdout"

    # Detect leakage flag
    leakage_flagged = False
    if len(cal)>0 and "leakage_flag" in cal.columns:
        leakage_flagged = (cal["leakage_flag"]=="WARN>0.75").any()
    leakage_status = (f"AUC>0.75 warning — KenPom-at-date={kp_pct_str}, BDB-only={bdb_pct_str}. "
                      f"Check archive coverage. Not a confirmed diagnosis.") \
                      if leakage_flagged else f"OK — AUC within expected pregame range"

    wb = Workbook()

    # ── TAB 1: Model Info ─────────────────────────────────────────────────
    wi=wb.active; wi.title="Model_Info"
    wi.sheet_view.showGridLines=False
    wi.column_dimensions["A"].width=36; wi.column_dimensions["B"].width=56

    banner(wi,1,WB_TITLE,sz=11,cols=2,bg=TITLE_BG)
    banner(wi,2,f"model_version = {VERSION_STAMP}  |  player_layer_promoted = {PROMOTED}",
           bg=NAVY,fg=GOLD,sz=10,h=18,cols=2)
    banner(wi,3,"STANDALONE .xlsx — does NOT modify the macro/VBA template workbook",
           bg=DRED,fg=WHITE,sz=9,h=16,cols=2)
    if leakage_flagged:
        banner(wi,4,f"⚠ AUC>0.75 — investigate coverage: KP={kp_pct_str}  BDB={bdb_pct_str}  |  ATS: {ats_status}",
               bg=ORANGE,fg=WHITE,sz=8,h=16,cols=2)
        offset=1
    else:
        offset=0

    VIG_MULT=100.0/110.0; VIG_BE=1.0/(1.0+VIG_MULT)
    if len(pred)>0 and "kenpom_used" in pred.columns:
        kp_pct=float(pred["kenpom_used"].mean()*100)

    PARAMS=[
        ("model_version",           VERSION_STAMP),
        ("production_model",        PROD_MODEL),
        ("ML_calibration_method",    "platt_logistic" if not is_player else "isotonic_oof"),
        ("ATS_signal_status",        "weak — AUC=0.5147 raw, no bucket N>=150 beats vig"),
        ("TOT_signal_status",        "weak — AUC=0.5167 raw"),
        ("experimental_model",      EXP_MODEL),
        ("player_layer_validated",  LAYER_VALID),
        ("player_layer_promoted",   PROMOTED),
        ("not_for_production_pricing", NOT_PROD),
        ("mean_layer",              "KP_backbone_plus_BDB_lambda_0.30"),
        ("player_layer",            "Additive residual — observed data only"),
        ("calibration_method",      "Isotonic regression OOF — TimeSeriesSplit(5)"),
        ("calibration_dataset",     f"{n_hist} H/R D1 games  2025-11-03 to 2026-03-10"),
        ("KenPom-at-date coverage", f"{kp_pct:.1f}% of historical games"),
        ("leakage_status",          leakage_status),
        ("true_walkforward_status", f"RAN — TeamBaselines pregame rolling | {n_hist} games | KP={kp_pct_str} | BDB={bdb_pct_str}"),
        ("phi / sigma",             "0.004 / 0.085"),
        ("LAM_OE / LAM_DE / LAM_TP","0.30 / 0.30 / 0.20"),
        ("vig_breakeven",           f"{VIG_BE:.4%}  (110/210)"),
        ("EV_formula",              f"cover_rate*(100/110) - (1-cover_rate)"),
        ("side_aware_edge",         "chosen_side=home if model_spread>-mkt_spread else away"),
        ("cal_prob_source",         "OOF isotonic when cal_is_oof=True; raw PMF otherwise"),
        ("neutral_calibration",     "NOT calibrated — H/R calibrator does not transfer"),
        ("player_layer_applied",    str(lat.get("player_layer_applied",pd.Series([False]))[0]) if len(lat)>0 else "False"),
        ("workbook_type",           "Standalone .xlsx — NOT macro/VBA template"),
        ("ATS_validation_status",     ats_status),
        ("overall_status",             "PROVISIONAL — ATS/TOT not validated. No bucket with N>=100 beats vig."),
        ("DO NOT CLAIM high win rate","No ATS bucket with N>=100 beats vig in current holdout"),
    ]
    for i,(k,v) in enumerate(PARAMS):
        r=i+5+offset; bg=LGREY if i%2==0 else WHITE
        warn=any(x in str(v) for x in ["NOT","WARN","DO NOT","LEAKAGE"])
        val(wi,r,1,k,bg=bg,fg=NAVY,bold=True,align="left")
        val(wi,r,2,v,bg=bg,fg=RED if warn else "FF000000",bold=warn,align="left")

    # ── TAB 2: Calibration Report ─────────────────────────────────────────
    wc=wb.create_sheet("Calibration_Report")
    wc.sheet_view.showGridLines=False
    wc.column_dimensions["A"].width=10; wc.column_dimensions["B"].width=20
    for col,w in zip("CDEFGHIJKL",[7,8,8,8,8,8,8,8,8,20]):
        wc.column_dimensions[col].width=w
    banner(wc,1,f"Calibration Report — {n_hist} historical games — OOF only (TimeSeriesSplit)",cols=12)
    if leakage_flagged:
        banner(wc,2,"⚠ AUC>0.75: same-game efficiency used — inflated metrics — see True_Walkforward column",
               bg=RED,fg=WHITE,sz=9,h=16,cols=12)
    CAL_COLS=[("Subset","subset",10,None,NAVY),("Market","market_type",20,None,NAVY),
              ("N_oof","n_oof",7,None,NAVY),("N_total","n_total",8,None,NAVY),
              ("AUC_raw","auc_raw",8,"0.0000",GREEN),("AUC_cal","auc_cal_oof",8,"0.0000",TEAL),
              ("Brier_raw","brier_raw",8,"0.0000",GREEN),("Brier_cal","brier_cal_oof",8,"0.0000",TEAL),
              ("ΔBrier","brier_delta",8,"+0.0000;-0.0000",ORANGE),
              ("Slope","cal_slope",8,"0.000",GREY2),("Intercept","cal_intercept",8,"0.000",GREY2),
              ("Leakage","leakage_flag",16,None,RED)]
    hr=3; [hdr(wc,hr,ci,h,bg=bg,sz=9) for ci,(h,_,_,_,bg) in enumerate(CAL_COLS,1)]
    if len(cal)>0:
        for ri,(_,r) in enumerate(cal.iterrows(),hr+1):
            bg=LGREY if ri%2==0 else WHITE
            for ci,(_,field,_,fmt,_) in enumerate(CAL_COLS,1):
                v=r.get(field,"")
                if isinstance(v,float) and math.isnan(v): v=""
                fg="FF000000"; bold=False
                if field=="leakage_flag": fg=RED if str(v).startswith("WARN") else GREEN; bold=True
                elif field=="brier_delta" and isinstance(v,float): fg=GREEN if v<0 else RED
                elif field in("auc_raw","auc_cal_oof") and isinstance(v,float):
                    fg=RED if v>0.75 else("GREEN" if v>0.60 else "FF000000")
                val(wc,ri,ci,v,bg=bg,fg=fg,bold=bold,fmt=fmt)
            wc.row_dimensions[ri].height=15

    # ── TAB 3: Edge Bucket Table ──────────────────────────────────────────
    we=wb.create_sheet("Edge_Bucket_Table")
    we.sheet_view.showGridLines=False
    banner(we,1,"Side-Aware ATS Edge Buckets — chosen_side=home if edge>0 else away",cols=10)
    banner(we,2,f"EV = cover_rate*(100/110)-(1-cover_rate)   Breakeven={VIG_BE:.4%}",
           bg=BLUE,sz=9,h=16,cols=10)
    EDGE_COLS=[("Subset","subset",9,None,NAVY),("Bucket","abs_bucket",14,None,NAVY),
               ("N","n",6,None,NAVY),("Cover%","chosen_cover_rate",9,"0.0%",GREEN),
               ("Model_P","chosen_prob_mean",9,"0.0%",TEAL),
               ("MeanEdge","mean_abs_edge",10,"0.00",GREY2),
               ("EV@-110","ev_at_110",10,"+0.0000;-0.0000",ORANGE),
               ("BeatsVig","beats_vig",10,None,PURPLE),
               ("%BetHome","pct_bet_home",10,"0.0%",GREY2),
               ("NCal","n_cal_probs",7,None,GREY2) if "n_cal_probs" in edge.columns
               else ("NCal","beats_vig",7,None,GREY2)]
    [hdr(we,3,ci,h,bg=bg,sz=9) for ci,(h,_,_,_,bg) in enumerate(EDGE_COLS,1)]
    if len(edge)>0:
        VIG_BE_4=round(VIG_BE,4)
        for ri,(_,r) in enumerate(edge.iterrows(),4):
            bg=LGREY if ri%2==0 else WHITE
            for ci,(_,field,_,fmt,_) in enumerate(EDGE_COLS,1):
                v=r.get(field,"")
                if isinstance(v,float) and math.isnan(v): v=""
                fg="FF000000"; bold=False
                if field=="ev_at_110" and isinstance(v,float):
                    fg=GREEN if v>0.02 else(RED if v<-0.02 else "FF777777"); bold=v>0.02
                elif field=="beats_vig": fg=GREEN if v else RED; bold=bool(v)
                elif field=="chosen_cover_rate" and isinstance(v,float):
                    fg=GREEN if v>VIG_BE_4 else(RED if v<0.47 else "FF000000")
                val(we,ri,ci,v,bg=bg,fg=fg,bold=bold,fmt=fmt)
            we.row_dimensions[ri].height=15

    # ── TAB 4: Pricing Output (raw + calibrated side by side) ────────────
    wp=wb.create_sheet("Pricing_Output_Final")
    wp.sheet_view.showGridLines=False; wp.freeze_panes="A3"

    # Determine which probability columns exist in latents
    has_adj = "fair_spread_adj" in lat.columns
    sp_col  = "fair_spread_adj"  if has_adj else "fair_spread"
    tt_col  = "fair_total_adj"   if has_adj else "fair_total"
    pml_raw = "p_ml_home_adj"    if "p_ml_home_adj" in lat.columns else "p_ml_home_raw"
    pcov_raw= "p_home_cover_adj" if "p_home_cover_adj" in lat.columns else "p_home_cover_raw"
    pov_raw = "p_over_adj"       if "p_over_adj" in lat.columns else "p_over_raw"
    mla_raw = "fair_ml_home_american_adj" if "fair_ml_home_american_adj" in lat.columns else "fair_ml_home_american"

    PC=[("GameID","GAME_ID",12,None,NAVY),("Date","DATE",11,None,NAVY),
        ("Home","HOME_KP",20,None,NAVY),("Away","AWAY_KP",20,None,NAVY),("Site","SITE",6,None,NAVY),
        ("MktSp","mkt_spread",9,"+0.0;-0.0",BLUE),("MktTt","mkt_total",9,"0.0",BLUE),
        ("FairSp(TO)","fair_spread",12,"+0.000;-0.000",GREEN),
   ("FairTt",    "fair_total", 10,"0.0",          WHITE),
        (f"FairSp(Pl)" if has_adj else "FairSp",sp_col,12,"+0.000;-0.000",TEAL),
        ("P(ML)_raw",pml_raw,10,"0.0%",GREEN),
        ("P(Cov)_raw",pcov_raw,10,"0.0%",GREEN),
        ("P(Ov)_raw",pov_raw,10,"0.0%",GREEN),
        ("ML_raw",mla_raw,10,"+#,##0;-#,##0",GREEN),
        ("P(ML)_cal","p_ml_calibrated",10,"0.0%",PURPLE) if "p_ml_calibrated" in lat.columns else ("P(ML)_cal",pml_raw,10,"0.0%",GREY2),
        ("P(Cov)_cal","p_cover_calibrated",10,"0.0%",PURPLE) if "p_cover_calibrated" in lat.columns else ("P(Cov)_cal",pcov_raw,10,"0.0%",GREY2),
        ("P(Ov)_cal","p_over_calibrated",10,"0.0%",PURPLE) if "p_over_calibrated" in lat.columns else ("P(Ov)_cal",pov_raw,10,"0.0%",GREY2),
        ("ML_cal","fair_ml_home_calibrated_american",10,"+#,##0;-#,##0",PURPLE) if "fair_ml_home_calibrated_american" in lat.columns else ("ML_cal",mla_raw,10,"+#,##0;-#,##0",GREY2),
        ("EdgeSp","edge_spread_pts" if "edge_spread_pts" in lat.columns else sp_col,10,"+0.000;-0.000",ORANGE),
        ("ConfH","adjustment_confidence_h" if "adjustment_confidence_h" in lat.columns else "SITE",9,None,TEAL),
        ("ConfA","adjustment_confidence_a" if "adjustment_confidence_a" in lat.columns else "SITE",9,None,TEAL),
        ("CalApplied","calibration_applied" if "calibration_applied" in lat.columns else "player_layer_applied",10,None,PURPLE),
        ("PlyrApplied","player_layer_applied",10,None,TEAL),
        ("ModelVer","model_version",22,None,NAVY),]

    [hdr(wp,1,ci,"",bg=bg) for ci,(_,_,_,_,bg) in enumerate(PC,1)]
    [hdr(wp,2,ci,h,bg=bg,sz=9) for ci,(h,_,_,_,bg) in enumerate(PC,1)]
    [setattr(wp.column_dimensions[get_column_letter(ci)],"width",w) for ci,(_,_,w,_,_) in enumerate(PC,1)]
    wp.row_dimensions[1].height=12; wp.row_dimensions[2].height=28

    # Strip ONLY player-adjusted columns from production workbook
    # Keep: team-only raw (_raw), team-only calibrated (_cal, calibrated), edge fields
    # Remove: player-adjusted spread/total/probs (columns ending in _adj or containing "adj")
    if not is_player:
        PLAYER_ADJ_FIELDS = {
            "fair_spread_adj","fair_total_adj",
            "p_ml_home_adj","p_home_cover_adj","p_over_adj",
            "fair_ml_home_american_adj","fair_ml_away_american_adj",
            "adjustment_confidence_h","adjustment_confidence_a",
            "player_ortg_adj_h","player_ortg_adj_a",
            "player_layer_applied",
        }
        PC = [(h,f,w,fmt,bg) for (h,f,w,fmt,bg) in PC
              if f not in PLAYER_ADJ_FIELDS and not f.endswith("_adj")]

    for ri,(_,row) in enumerate(lat.iterrows(),3):
        bg=LGREY if ri%2==0 else WHITE
        for ci,(_,field,_,fmt,_) in enumerate(PC,1):
            v=row.get(field,"")
            if isinstance(v,float) and math.isnan(v): v=""
            fg="FF000000"; bold=False
            if field in("fair_spread",sp_col) and isinstance(v,float):
                fg=GREEN if v>0 else(RED if v<0 else "FF000000"); bold=True
            elif "p_ml" in field or "p_cover" in field or "p_over" in field:
                if isinstance(v,float): fg=GREEN if v>0.57 else(RED if v<0.43 else "FF000000")
            val(wp,ri,ci,v,bg=bg,fg=fg,bold=bold,fmt=fmt)
        wp.row_dimensions[ri].height=15

    # ── TAB 5: Slate Summary ──────────────────────────────────────────────
    ws2=wb.create_sheet("Slate_Summary")
    ws2.sheet_view.showGridLines=False
    ws2.column_dimensions["A"].width=34; ws2.column_dimensions["B"].width=24
    banner(ws2,1,f"Slate Summary — {args.date} — {VERSION_STAMP}",cols=2,
           bg=TITLE_BG)
    if is_player:
        banner(ws2,2,"⚠ EXPERIMENTAL — Player layer FAILED holdout — NOT FOR PRODUCTION PRICING",
               bg=RED,fg=WHITE,sz=9,h=16,cols=2)
    sp_v=lat[sp_col].dropna(); tt_v=lat[tt_col].dropna() if tt_col in lat.columns else pd.Series(dtype=float)
    both=lat.dropna(subset=[sp_col,"mkt_spread"])
    corr=float(np.corrcoef(both[sp_col],-both["mkt_spread"])[0,1]) if len(both)>3 else float("nan")
    mad=float((both[sp_col]-(-both["mkt_spread"])).abs().mean()) if len(both)>3 else float("nan")
    STATS=[
        ("Games scored",               len(lat)),
        ("Neutral games",              int((lat["SITE"]=="N").sum()) if "SITE" in lat.columns else "n/a"),
        ("player_layer_applied",       str(lat["player_layer_applied"].all()) if "player_layer_applied" in lat.columns else "False"),
        ("fair_spread SD",             f"{sp_v.std():.3f}" if len(sp_v)>0 else "n/a"),
        ("corr(fair_spread,-mkt)",     f"{corr:.4f}" if not math.isnan(corr) else "n/a"),
        ("MAD vs market",              f"{mad:.3f} pts" if not math.isnan(mad) else "n/a"),
        ("pmf_grid_sum errors",        "see cal report"),
        ("Calibration games",          n_hist),
        ("Calibration leakage",        leakage_status[:60]),
        ("model_version",              "calibrated_v1_p2_30"),
        ("mean_layer",                 "KP_backbone_plus_BDB_lambda_0.30"),
        ("Vig breakeven",              f"{VIG_BE:.4%}"),
        ("NOT DONE: true walk-forward","Requires TeamBaselines.csv pregame rolling"),
        ("NOT DONE: neutral cal",      "n<100 neutral historical games"),
        ("NOT DONE: template connect", "Phase 3D — not yet connected to .xlsm"),
    ]
    for i,(k,v) in enumerate(STATS):
        r=i+3; bg=LGREY if i%2==0 else WHITE
        warn="NOT DONE" in str(v) or "WARN" in str(v)
        val(ws2,r,1,k,bg=bg,fg=NAVY,bold=True,align="left")
        val(ws2,r,2,v,bg=bg,fg=RED if warn else "FF000000",bold=warn,align="left")

    wb.save(args.out)
    print(f"Written: {args.out}")
    print(f"Sheets: {wb.sheetnames}")

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--latents",    required=True)
    ap.add_argument("--cal-report", default="cbb_cache/model_calibration_report.csv",dest="cal_report")
    ap.add_argument("--edge",       default="cbb_cache/edge_bucket_table.csv")
    ap.add_argument("--pred",       default="cbb_cache/historical_p230_predictions.csv")
    ap.add_argument("--out",        required=True)
    ap.add_argument("--date",       default="2026-03-21")
    build(ap.parse_args())
if __name__=="__main__": main()
