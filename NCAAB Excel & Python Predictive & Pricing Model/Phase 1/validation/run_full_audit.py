
import openpyxl, pandas as pd, numpy as np
from sklearn.metrics import brier_score_loss, roc_auc_score, log_loss
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression as PlattLR, LinearRegression
from sklearn.model_selection import TimeSeriesSplit
from sklearn.calibration import calibration_curve

VIG_MULT=100/110; VIG_BE=1/(1+VIG_MULT)
def ats_ev(cr): return round(cr*VIG_MULT-(1-cr),4)

def cal_slope_intercept(yt, yp, n_bins=10):
    nb=min(n_bins,len(yt)//50)
    if nb<3: return float("nan"),float("nan")
    pt,pp2=calibration_curve(yt,yp,n_bins=nb)
    lm=LinearRegression().fit(pp2.reshape(-1,1),pt)
    return round(float(lm.coef_[0]),3),round(float(lm.intercept_),3)

def full_metrics(df, y, p, platt=False):
    s=df.dropna(subset=[y,p]).copy()
    s=s[s[p].between(0.01,0.99)].sort_values("DATE").reset_index(drop=True)
    if len(s)<50: return None
    yt=s[y].values.astype(float); yp=s[p].values
    ar=round(roc_auc_score(yt,yp),4)
    br=round(brier_score_loss(yt,yp),4)
    lr=round(log_loss(yt,yp),4)
    sl_r,ic_r=cal_slope_intercept(yt,yp)
    tss=TimeSeriesSplit(n_splits=5); op=[]; oc=[]
    for tr,te in tss.split(yt):
        if len(tr)<20: continue
        if platt:
            m=PlattLR(C=1e6,solver="lbfgs"); m.fit(yp[tr].reshape(-1,1),yt[tr])
            cp=m.predict_proba(yp[te].reshape(-1,1))[:,1]
        else:
            m=IsotonicRegression(out_of_bounds="clip"); m.fit(yp[tr],yt[tr]); cp=m.predict(yp[te])
        op.extend(te); oc.extend(cp.tolist())
    op=np.array(op); oc=np.array(oc)
    ac=round(roc_auc_score(yt[op],oc),4)
    bc=round(brier_score_loss(yt[op],oc),4)
    lc=round(log_loss(yt[op],np.clip(oc,1e-6,1-1e-6)),4)
    sl_c,ic_c=cal_slope_intercept(yt[op],oc)
    return {"n":len(s),"n_oof":len(op),"ar":ar,"br":br,"lr":lr,
            "ac":ac,"bc":bc,"lc":lc,
            "sl_r":sl_r,"ic_r":ic_r,"sl_c":sl_c,"ic_c":ic_c}

# Workbook audit
wb=openpyxl.load_workbook("outputs/ncaab_market_maker_2026-03-23_PRODUCTION.xlsx")
mi=wb["Model_Info"]
info={}
for row in mi.iter_rows(min_row=5,values_only=True):
    if row[0] and row[1]: info[str(row[0]).strip()]=str(row[1]).strip()

REQUIRED={
    "production_model":"team_only_v1_p2_30",
    "ML_calibration_method":"platt_logistic",
    "ATS_signal_status":"weak",
    "TOT_signal_status":"weak",
    "player_layer_validated":"False",
    "player_layer_promoted":"False",
    "not_for_production_pricing":"False",
}
print("="*65)
print("MODEL_INFO AUDIT:")
for k,exp in REQUIRED.items():
    v=info.get(k,"MISSING")
    ok="OK" if v!="MISSING" and exp.lower() in v.lower() else "MISSING"
    print(f"  [{ok}] {k} = {v}")

# Column presence audit
po=wb["Pricing_Output_Final"]
hdr=[str(c.value).strip() if c.value else "" for c in list(po.iter_rows(min_row=2,max_row=2))[0]]
hdr_set=set(hdr)

MUST_HAVE=["MktSp","MktTt","FairSp","FairTt","ML_raw","ML_cal",
           "P(ML)_raw","P(Cov)_raw","P(Ov)_raw",
           "P(ML)_cal","P(Cov)_cal","P(Ov)_cal",
           "EdgeSp","CalApplied","PlyrApplied","ModelVer"]
MUST_NOT=["fair_spread_adj","fair_total_adj","p_ml_home_adj",
          "p_home_cover_adj","p_over_adj","player_ortg_adj_h","player_ortg_adj_a"]

print("\nCOLUMN PRESENCE AUDIT (Pricing_Output_Final):")
print(f"  All columns: {[c for c in hdr if c]}")
print("\n  MUST HAVE:")
for c in MUST_HAVE:
    print(f"    [{'OK' if c in hdr_set else 'MISSING'}] {c}")
print("\n  MUST NOT HAVE (player-adj):")
for c in MUST_NOT:
    print(f"    [{'FAIL - PRESENT' if c in hdr_set else 'OK - absent'}] {c}")

# Validation table
preds=pd.read_csv("cbb_cache/historical_p230_predictions.csv",parse_dates=["DATE"])
hr=preds[preds["VENUE"]=="H/R"].sort_values("DATE").reset_index(drop=True)
neut=preds[preds["VENUE"]!="H/R"]
kp=int(preds["kenpom_used"].sum()) if "kenpom_used" in preds.columns else "n/a"

print("\n" + "="*65)
print(f"VALIDATION TABLE (min_games=0)")
print(f"  Total={len(preds)}  H/R={len(hr)}  Neutral={len(neut)}")
print(f"  KenPom={kp} ({100*kp/len(preds):.1f}%)  BDB={len(preds)-kp}")

ml =full_metrics(hr,"home_win","p_ml_home",platt=True)
ats=full_metrics(hr,"home_covered","p_home_cover")
tot=full_metrics(hr,"over","p_over")

methods={"ML (Platt)":"platt_logistic","ATS (isotonic)":"isotonic_oof","TOT (isotonic)":"isotonic_oof"}
print(f"\n  {'Market':<16} {'N':>5} {'N_oof':>5} {'AUC_r':>7} {'AUC_c':>7} {'Br_r':>7} {'Br_c':>7} {'LL_r':>7} {'LL_c':>7} {'Sl_r':>7} {'Ic_r':>7} {'Sl_c':>7} {'Ic_c':>7} {'Method'}")
print(f"  {'-'*110}")
for nm,met,m in [("ML (Platt)","platt_logistic",ml),("ATS (isotonic)","isotonic_oof",ats),("TOT (isotonic)","isotonic_oof",tot)]:
    if m:
        print(f"  {nm:<16} {m['n']:>5} {m['n_oof']:>5} {m['ar']:>7.4f} {m['ac']:>7.4f} {m['br']:>7.4f} {m['bc']:>7.4f} {m['lr']:>7.4f} {m['lc']:>7.4f} {m['sl_r']:>7.3f} {m['ic_r']:>7.3f} {m['sl_c']:>7.3f} {m['ic_c']:>7.3f} {met}")

# Edge buckets
p2=hr.dropna(subset=["mkt_spread","home_covered","p_home_cover"]).copy()
p2["home_covered"]=p2["home_covered"].astype(float)
p2["edge"]=p2["fair_spread"]-(-p2["mkt_spread"])
p2["chosen"]=np.where(p2["edge"]>0,p2["home_covered"],1-p2["home_covered"])
p2["abs_edge"]=p2["edge"].abs()
p2["bucket"]=pd.cut(p2["abs_edge"],bins=[0,1.5,3,5,99],labels=["0-1.5","1.5-3","3-5",">5"])
bkt=p2.groupby("bucket",observed=True).agg(n=("chosen","count"),cr=("chosen","mean")).reset_index()
bkt["ev"]=bkt["cr"].apply(ats_ev); bkt["vig"]=bkt["cr"]>VIG_BE
print(f"\n  ATS EDGE BUCKETS (H/R, breakeven=52.3810%):")
print(f"  {'Bucket':<10} {'N':>5} {'CovR':>7} {'EV@-110':>9} {'Beats vig':>10} {'N>=150':>7}")
for _,r in bkt.iterrows():
    print(f"  {str(r['bucket']):<10} {r['n']:>5} {r['cr']:>7.3f} {r['ev']:>9.4f} {'YES  ✓' if r['vig'] else 'no':>10} {'YES' if r['n']>=150 else 'no':>7}")

print("\n" + "="*65)
print("PRODUCTION WORKBOOK: outputs/ncaab_market_maker_2026-03-23_PRODUCTION.xlsx")
print("EXPERIMENTAL WORKBOOK: outputs/ncaab_market_maker_2026-03-23_EXPERIMENTAL_player.xlsx")
print("VERDICT: provisional pricing workbook only")
print("  ATS AUC=0.5127 TOT AUC=0.5157 — weak signal")
print("  Player layer experimental only — failed holdout")
print("  Not a validated high-win-rate betting model")
