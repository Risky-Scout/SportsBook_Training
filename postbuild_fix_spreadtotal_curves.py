#!/usr/bin/env python3
from __future__ import annotations
import argparse
from openpyxl import load_workbook

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--row_start", type=int, default=11)
    ap.add_argument("--row_end", type=int, default=200)
    args = ap.parse_args()

    wb = load_workbook(args.workbook)
    if "SpreadTotal" not in wb.sheetnames:
        raise SystemExit(f"{args.workbook} missing SpreadTotal sheet")
    ws = wb["SpreadTotal"]

    for r in range(args.row_start, args.row_end + 1):
        ws[f"C{r}"] = f"=NORMDIST(A{r},-Inputs!$B$18,Inputs!$B$16,FALSE)"
        ws[f"D{r}"] = f"=NORMDIST(A{r},Inputs!$B$14,Inputs!$B$16,FALSE)"
        ws[f"H{r}"] = f"=NORMDIST(F{r},Inputs!$B$19,Inputs!$B$17,FALSE)"
        ws[f"I{r}"] = f"=NORMDIST(F{r},Inputs!$B$15,Inputs!$B$17,FALSE)"

    wb.save(args.workbook)
    print(f"Patched SpreadTotal curves in {args.workbook}")

if __name__ == "__main__":
    main()
