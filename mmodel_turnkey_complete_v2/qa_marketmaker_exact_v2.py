#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path
from openpyxl import load_workbook


def is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def main():
    ap = argparse.ArgumentParser(description="Exact-workbook QA for NCAAB market maker V2 exact workbook")
    ap.add_argument("--workbook", required=True)
    args = ap.parse_args()

    path = Path(args.workbook)
    wb = load_workbook(path, data_only=False)

    issues = []

    required = [
        "MarketMaker_Dashboard",
        "GamePick_ModeA",
        "Inputs",
        "MarketMaker_Board",
        "SpreadTotal",
        "GameInputs",
        "BlendedRatings",
        "ExactPMF_Summary",
        "ExactPMF_Margin",
        "ExactPMF_Total",
    ]
    for s in required:
        if s not in wb.sheetnames:
            issues.append(f"Missing required sheet: {s}")

    if issues:
        print("QA FAILED")
        for x in issues:
            print(" -", x)
        raise SystemExit(1)

    dash = wb["MarketMaker_Dashboard"]
    inputs = wb["Inputs"]
    board = wb["MarketMaker_Board"]
    spread = wb["SpreadTotal"]
    summary = wb["ExactPMF_Summary"]
    gp = wb["GamePick_ModeA"]

    for cell in ["B2", "E2", "B4", "B5", "B6", "E4", "E5", "E6", "H4", "H5", "H6", "K4", "K5", "K6"]:
        if not is_formula(dash[cell].value):
            issues.append(f"Dashboard {cell} should be formula-driven")

    for cell in ["B3","B4","B5","B6","B7","B8","B9","B12","B13","B14","B15","B18","B19","B20","B21","B22","B23","B24","B25","B26","B27","B28","B29","B30"]:
        if not is_formula(inputs[cell].value):
            issues.append(f"Inputs {cell} should be formula-driven")

    header_expect = {
        "F1": "Market Home Line",
        "G1": "Fair Home Margin",
        "H1": "Home-Side Edge",
        "I1": "Market Total",
        "J1": "Fair Total",
        "K1": "Total Edge",
        "T1": "Fair Home Line",
    }
    for cell, text in header_expect.items():
        val = board[cell].value
        if str(val).strip() != text:
            issues.append(f"Board header {cell} expected '{text}' but found '{val}'")

    if not is_formula(spread["B11"].value):
        issues.append("SpreadTotal!B11 should be formula-driven exact margin PMF")
    if not is_formula(spread["G11"].value):
        issues.append("SpreadTotal!G11 should be formula-driven exact total PMF")

    if summary.max_row < 2:
        issues.append("ExactPMF_Summary has no data rows")

    try:
        sel = int(gp["B2"].value)
    except Exception:
        sel = None
        issues.append("GamePick_ModeA!B2 is not an integer selector")

    if sel is not None:
        for cell in ["B7","B8","B9","B12","B13","B14","B15","B20","B21","B22"]:
            if "ExactPMF_Summary" not in str(inputs[cell].value):
                issues.append(f"Inputs {cell} does not reference ExactPMF_Summary")

    if issues:
        print("QA FAILED")
        for x in issues:
            print(" -", x)
        raise SystemExit(1)

    print("QA PASSED (EXACT MODE)")


if __name__ == "__main__":
    main()
