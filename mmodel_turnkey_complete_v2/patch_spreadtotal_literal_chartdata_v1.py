#!/usr/bin/env python3
from __future__ import annotations

import argparse
import math
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.shapes import GraphicalProperties


def to_float(x):
    try:
        if x is None or x == "":
            return None
        return float(x)
    except Exception:
        return None


def build_scatter_chart(ws, title: str, x_title: str, y_title: str,
                        x_col: int, y_col: int, market_col: int, model_col: int,
                        min_row: int, max_row: int, width: float = 10.5, height: float = 5.2):
    chart = ScatterChart()
    chart.title = title
    chart.style = 2
    chart.width = width
    chart.height = height
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.y_axis.number_format = "0.0%"
    chart.legend.position = "r"

    x = Reference(ws, min_col=x_col, min_row=min_row, max_row=max_row)
    y = Reference(ws, min_col=y_col, min_row=min_row, max_row=max_row)
    mkt = Reference(ws, min_col=market_col, min_row=min_row, max_row=max_row)
    mdl = Reference(ws, min_col=model_col, min_row=min_row, max_row=max_row)

    s1 = Series(y, x, title="Exact PMF")
    s1.marker.symbol = "none"
    s1.graphicalProperties = GraphicalProperties()
    s1.graphicalProperties.line.width = 22000
    chart.series.append(s1)

    s2 = Series(mkt, x, title="Market")
    s2.graphicalProperties = GraphicalProperties()
    s2.graphicalProperties.line.noFill = True
    s2.marker.symbol = "diamond"
    s2.marker.size = 9
    chart.series.append(s2)

    s3 = Series(mdl, x, title="Model")
    s3.graphicalProperties = GraphicalProperties()
    s3.graphicalProperties.line.noFill = True
    s3.marker.symbol = "triangle"
    s3.marker.size = 10
    chart.series.append(s3)

    return chart


def rows_from_exact_sheet(ws_exact, cutoff: str, home_team: str, away_team: str):
    out = []
    # Expected columns: A Cutoff, B Home Team, C Away Team, D Margin/Total, E PMF
    for r in range(2, ws_exact.max_row + 1):
        if ws_exact.cell(r, 1).value == cutoff and ws_exact.cell(r, 2).value == home_team and ws_exact.cell(r, 3).value == away_team:
            x = to_float(ws_exact.cell(r, 4).value)
            p = to_float(ws_exact.cell(r, 5).value)
            if x is not None and p is not None:
                out.append((x, p))
    return out


def write_chart_block(ws, start_col: int, start_row: int, header_x: str, data: list[tuple[float, float]],
                      center_value: float, model_value: float, window_half_width: int):
    # Hidden literal numeric data block
    headers = [header_x, "Dist Prob", "Market Marker", "Model Marker"]
    for i, h in enumerate(headers):
        ws.cell(start_row, start_col + i, h)

    # Build integer grid around center
    x_start = int(round(center_value - window_half_width))
    x_end = int(round(center_value + window_half_width))
    grid = list(range(x_start, x_end + 1))
    prob_map = {int(round(x)): p for x, p in data}

    for idx, x in enumerate(grid, start=1):
        row = start_row + idx
        p = prob_map.get(int(round(x)), 0.0)
        ws.cell(row, start_col, x)
        ws.cell(row, start_col + 1, p)
        ws.cell(row, start_col + 2, p if x == int(round(center_value)) else None)
        ws.cell(row, start_col + 3, p if x == int(round(model_value)) else None)
        ws.cell(row, start_col + 1).number_format = "0.0%"
        ws.cell(row, start_col + 2).number_format = "0.0%"
        ws.cell(row, start_col + 3).number_format = "0.0%"

    return start_row + 1, start_row + len(grid)


def main():
    ap = argparse.ArgumentParser(description="Rebuild SpreadTotal charts from literal numeric chart data.")
    ap.add_argument("--workbook", required=True)
    args = ap.parse_args()

    path = Path(args.workbook)

    # Load twice: data_only workbook for cached numeric values, normal workbook for writing
    wb_vals = load_workbook(path, data_only=True)
    wb = load_workbook(path)

    inputs_vals = wb_vals["Inputs"]
    spread_vals = wb_vals["SpreadTotal"]
    exact_margin_vals = wb_vals["ExactPMF_Margin"]
    exact_total_vals = wb_vals["ExactPMF_Total"]

    ws = wb["SpreadTotal"]

    cutoff = str(inputs_vals["B3"].value)
    home_team = str(inputs_vals["B4"].value)
    away_team = str(inputs_vals["B5"].value)

    fair_home_margin = to_float(inputs_vals["B14"].value)
    fair_total = to_float(inputs_vals["B15"].value)
    margin_sd = to_float(inputs_vals["B16"].value)
    total_sd = to_float(inputs_vals["B17"].value)
    market_home_line = to_float(inputs_vals["B18"].value)
    market_total = to_float(inputs_vals["B19"].value)

    if None in [fair_home_margin, fair_total, margin_sd, total_sd, market_home_line, market_total]:
        raise SystemExit("Could not read cached numeric values from Inputs sheet. Open/save workbook in Excel first, then rerun.")

    margin_data = rows_from_exact_sheet(exact_margin_vals, cutoff, home_team, away_team)
    total_data = rows_from_exact_sheet(exact_total_vals, cutoff, home_team, away_team)

    if not margin_data:
        raise SystemExit("No matching rows found in ExactPMF_Margin for selected game.")
    if not total_data:
        raise SystemExit("No matching rows found in ExactPMF_Total for selected game.")

    # Use hidden helper columns AA:AH only; leave visible sheet untouched.
    for col in ["AA", "AB", "AC", "AD", "AF", "AG", "AH", "AI"]:
        ws.column_dimensions[col].hidden = True

    # Margin block in AA:AD
    margin_center = -market_home_line
    margin_half = max(6, int(round(3 * abs(margin_sd))))
    m_min_row, m_max_row = write_chart_block(
        ws,
        start_col=27,  # AA
        start_row=10,
        header_x="Home Margin",
        data=margin_data,
        center_value=margin_center,
        model_value=fair_home_margin,
        window_half_width=margin_half,
    )

    # Total block in AF:AI
    total_half = max(8, int(round(3 * abs(total_sd))))
    t_min_row, t_max_row = write_chart_block(
        ws,
        start_col=32,  # AF
        start_row=10,
        header_x="Game Total",
        data=total_data,
        center_value=market_total,
        model_value=fair_total,
        window_half_width=total_half,
    )

    # Remove existing charts and rebuild from literal numeric data
    ws._charts = []

    margin_chart = build_scatter_chart(
        ws,
        title="Home Margin Exact PMF | Market vs Model",
        x_title="Home margin (points)",
        y_title="Probability mass",
        x_col=27, y_col=28, market_col=29, model_col=30,
        min_row=m_min_row, max_row=m_max_row,
        width=10.6, height=5.2
    )
    total_chart = build_scatter_chart(
        ws,
        title="Game Total Exact PMF | Market vs Model",
        x_title="Game total (points)",
        y_title="Probability mass",
        x_col=32, y_col=33, market_col=34, model_col=35,
        min_row=t_min_row, max_row=t_max_row,
        width=10.6, height=5.2
    )

    ws.add_chart(margin_chart, "J28")
    ws.add_chart(total_chart, "J52")

    wb.save(path)
    print(f"Rebuilt SpreadTotal charts from literal numeric chart data in {path}")


if __name__ == "__main__":
    main()
