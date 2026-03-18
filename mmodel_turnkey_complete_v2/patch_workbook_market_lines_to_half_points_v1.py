#!/usr/bin/env python3
from openpyxl import load_workbook
from pathlib import Path
import argparse
import math

def snap_half(x):
    x = float(x)
    return math.copysign(math.floor(abs(x) * 2 + 0.5) / 2, x)

ap = argparse.ArgumentParser()
ap.add_argument("--workbook", required=True)
args = ap.parse_args()

path = Path(args.workbook)
wb = load_workbook(path)

# 1) Snap the internal GameInputs sheet values (this is what the board reads)
if "GameInputs" not in wb.sheetnames:
    raise SystemExit("Workbook missing internal GameInputs sheet")

g = wb["GameInputs"]
changed = 0
for r in range(2, g.max_row + 1):
    for c in (5, 6):  # E=Home spread line (input), F=Game total line (input)
        v = g.cell(row=r, column=c).value
        try:
            new_v = snap_half(v)
            if v != new_v:
                g.cell(row=r, column=c).value = new_v
                changed += 1
        except Exception:
            pass

# 2) Force the visible board to read snapped half-point values
if "MarketMaker_Board" in wb.sheetnames:
    ws = wb["MarketMaker_Board"]
    max_r = max(ws.max_row, 200)
    for r in range(2, max_r + 1):
        ws[f"F{r}"] = f'=IFERROR(ROUND(INDEX(GameInputs!$E:$E,$A{r}+1)*2,0)/2,"")'
        ws[f"I{r}"] = f'=IFERROR(ROUND(INDEX(GameInputs!$F:$F,$A{r}+1)*2,0)/2,"")'

# 3) Force Excel full recalc on open
try:
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
except Exception:
    pass

wb.save(path)
print(f"Patched workbook market lines to 0.5 grid in {path}")
print(f"Internal GameInputs cells changed: {changed}")
