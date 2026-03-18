#!/usr/bin/env python3
from __future__ import annotations
import argparse
from openpyxl import load_workbook

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True)
    args = ap.parse_args()

    wb = load_workbook(args.workbook)
    if "MarketMaker_Dashboard" not in wb.sheetnames:
        raise SystemExit(f"{args.workbook} missing MarketMaker_Dashboard sheet")
    dash = wb["MarketMaker_Dashboard"]

    dash["B2"] = '=IFERROR(INDEX(ExactPMF_Summary!$C:$C,GamePick_ModeA!$B$2+1)&" @ "&INDEX(ExactPMF_Summary!$B:$B,GamePick_ModeA!$B$2+1),"")'
    dash["E2"] = '=GamePick_ModeA!$B$2'
    dash["B4"] = '=IFERROR(INDEX(ExactPMF_Summary!$E:$E,GamePick_ModeA!$B$2+1),"")'
    dash["B5"] = '=IFERROR(INDEX(ExactPMF_Summary!$L:$L,GamePick_ModeA!$B$2+1),"")'
    dash["B6"] = '=IF(AND(ISNUMBER(B4),ISNUMBER(B5)),B5+B4,"")'
    dash["E4"] = '=IFERROR(INDEX(ExactPMF_Summary!$F:$F,GamePick_ModeA!$B$2+1),"")'
    dash["E5"] = '=IFERROR(INDEX(ExactPMF_Summary!$M:$M,GamePick_ModeA!$B$2+1),"")'
    dash["E6"] = '=IF(AND(ISNUMBER(E4),ISNUMBER(E5)),E5-E4,"")'
    dash["H4"] = '=IFERROR(INDEX(ExactPMF_Summary!$N:$N,GamePick_ModeA!$B$2+1),"")'
    dash["H5"] = '=IFERROR(INDEX(ExactPMF_Summary!$O:$O,GamePick_ModeA!$B$2+1),"")'
    dash["H6"] = '=IFERROR(INDEX(ExactPMF_Summary!$P:$P,GamePick_ModeA!$B$2+1),"")'
    dash["K4"] = '=IFERROR(INDEX(ExactPMF_Summary!$T:$T,GamePick_ModeA!$B$2+1),"")'
    dash["K5"] = '=IFERROR(INDEX(ExactPMF_Summary!$U:$U,GamePick_ModeA!$B$2+1),"")'
    dash["K6"] = '=IFERROR(INDEX(ExactPMF_Summary!$V:$V,GamePick_ModeA!$B$2+1),"")'

    wb.save(args.workbook)
    print(f"Patched dashboard selected-game block in {args.workbook}")

if __name__ == "__main__":
    main()
