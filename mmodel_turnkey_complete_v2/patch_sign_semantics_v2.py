#!/usr/bin/env python3
from __future__ import annotations
import argparse
import shutil
from pathlib import Path
import openpyxl

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True)
    args = ap.parse_args()

    path = Path(args.workbook)
    if not path.exists():
        raise SystemExit(f"Workbook not found: {path}")

    bak = path.with_suffix(path.suffix + ".signlabels.bak")
    shutil.copy2(path, bak)

    wb = openpyxl.load_workbook(path)

    if "Inputs" in wb.sheetnames:
        ws = wb["Inputs"]
        labels = {
            "A14": "Fair Home Margin",
            "A15": "Fair Total",
            "A18": "Market Home Line",
            "A23": "Home-Side Edge",
            "A24": "Total Edge",
            "A27": "Fair Cover Decimal",
            "A28": "Fair Cover American",
            "A29": "Fair Over Decimal",
            "A30": "Fair Over American",
        }
        for cell, value in labels.items():
            ws[cell] = value
        ws["D14"] = "Fair Home Line"
        ws["E14"] = '=-$B$14'
        ws["E14"].number_format = "0.00"
        if ws["D15"].value in (None, ""):
            ws["D15"] = "Convention"
            ws["E15"] = "Home line = negative favorite / positive dog"

    if "MarketMaker_Dashboard" in wb.sheetnames:
        ws = wb["MarketMaker_Dashboard"]
        relabel = {
            "A4": "Market Home Line",
            "A5": "Fair Home Margin",
            "A6": "Home-Side Edge",
            "D4": "Market Total",
            "D5": "Fair Total",
            "D6": "Total Edge",
            "G4": "Home Win %",
            "G5": "Home Cover %",
            "G6": "Over %",
            "J4": "Fair ML US",
            "J5": "Fair Cover US",
            "J6": "Fair Over US",
            "L3": "Chart guide",
            "L4": "Dist Prob = exact PMF probability mass",
            "L5": "Market Home Line uses betting sign",
            "L6": "Fair Home Margin > 0 means home win",
        }
        for cell, value in relabel.items():
            ws[cell] = value

    if "MarketMaker_Board" in wb.sheetnames:
        ws = wb["MarketMaker_Board"]
        header_map = {
            "F1": "Market Home Line",
            "G1": "Fair Home Margin",
            "H1": "Home-Side Edge",
            "I1": "Market Total",
            "J1": "Fair Total",
            "K1": "Total Edge",
            "L1": "Home Win %",
            "M1": "Home Cover %",
            "N1": "Over %",
            "O1": "Fair ML Dec",
            "P1": "Fair ML US",
            "Q1": "Fair Cover Dec",
            "R1": "Fair Cover US",
            "S1": "Fair Over Dec",
            "T1": "Fair Over US",
            "U1": "Fair Home Line",
        }
        for cell, value in header_map.items():
            ws[cell] = value

        for r in range(2, ws.max_row + 1):
            if ws[f"B{r}"].value not in (None, "", 0, "0"):
                ws[f"U{r}"] = f'=IF(ISNUMBER(G{r}),-G{r},"")'
                ws[f"U{r}"].number_format = "0.00"

    if "SpreadTotal" in wb.sheetnames:
        ws = wb["SpreadTotal"]
        note_cells = {
            "N2": "Sign convention",
            "N3": "Fair Home Margin > 0 means home expected to win.",
            "N4": "Fair Home Line = - Fair Home Margin.",
            "N5": "Market Home Line: negative favorite / positive dog.",
        }
        for cell, value in note_cells.items():
            ws[cell] = value

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(path)

    print(f"Patched sign/semantic labels in: {path}")
    print(f"Backup created at: {bak}")

if __name__ == "__main__":
    main()
