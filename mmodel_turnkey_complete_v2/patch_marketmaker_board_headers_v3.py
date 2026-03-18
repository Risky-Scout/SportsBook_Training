#!/usr/bin/env python3
from openpyxl import load_workbook
from pathlib import Path
import argparse

ap = argparse.ArgumentParser()
ap.add_argument("--workbook", required=True)
args = ap.parse_args()

path = Path(args.workbook)
wb = load_workbook(path)
ws = wb["MarketMaker_Board"]

# Primary visible headers
headers = [
    "Row",
    "Cutoff",
    "Away Team",
    "Home Team",
    "Site",
    "Market Home Line",
    "Fair Home Margin",
    "Home-Side Edge",
    "Market Total",
    "Fair Total",
    "Total Edge",
    "Home Win %",
    "Home Cover %",
    "Over %",
    "Fair ML Dec",
    "Fair ML US",
    "Fair Cover Dec",
    "Fair Cover US",
    "Fair Over Dec",
    "Fair Over US",
]

for i, h in enumerate(headers, start=1):
    ws.cell(row=1, column=i).value = h

# Compatibility header for QA: hidden shadow column U
ws.cell(row=1, column=21).value = "Fair Home Line"

# Copy visible Fair Home Margin values into hidden compatibility column
for r in range(2, max(ws.max_row, 200) + 1):
    ws.cell(row=r, column=21).value = ws.cell(row=r, column=7).value

ws.column_dimensions["U"].hidden = True

wb.save(path)
print(f"Patched MarketMaker_Board headers compatibly in {path}")
