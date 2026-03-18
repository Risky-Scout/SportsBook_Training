#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path
from openpyxl import load_workbook


def is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def find_header_map(ws, header_row: int = 1) -> dict[str, str]:
    out = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        text = str(cell.value).strip()
        if text:
            out[text] = cell.coordinate
    return out


def main():
    ap = argparse.ArgumentParser(description="Exact-workbook QA for NCAAB market maker V2 exact workbook")
    ap.add_argument("--workbook", required=True)
    args = ap.parse_args()

    wb = load_workbook(Path(args.workbook), data_only=False)
    issues = []

    required = [
        "MarketMaker_Dashboard",
        "GamePick_ModeA",
        "Inputs",
        "MarketMaker_Board",
        "SpreadTotal",
        "GameInputs",
        "BlendedRatings",
        "ExactPMF_Summary",
        "ExactPMF_Margin",
        "ExactPMF_Total",
    ]
    for s in required:
        if s not in wb.sheetnames:
            issues.append(f"Missing required sheet: {s}")

    if issues:
        print("QA FAILED")
        for x in issues:
            print(" -", x)
        raise SystemExit(1)

    dash = wb["MarketMaker_Dashboard"]
    inputs = wb["Inputs"]
    board = wb["MarketMaker_Board"]
    spread = wb["SpreadTotal"]
    summary = wb["ExactPMF_Summary"]
    gp = wb["GamePick_ModeA"]

    # Dashboard selected block should still be formula-driven
    for cell in ["B2", "E2", "B4", "B5", "B6", "E4", "E5", "E6", "H4", "H5", "H6", "K4", "K5", "K6"]:
        if not is_formula(dash[cell].value):
            issues.append(f"Dashboard {cell} should be formula-driven")

    # Inputs selected-game values should be formula-driven
    for cell in ["B3","B4","B5","B6","B7","B8","B9","B12","B13","B14","B15","B18","B19","B20","B21","B22","B23","B24","B25","B26","B27","B28","B29","B30"]:
        if not is_formula(inputs[cell].value):
            issues.append(f"Inputs {cell} should be formula-driven")

    # Inputs exact references
    for cell in ["B7","B8","B9","B12","B13","B14","B15","B20","B21","B22"]:
        if "ExactPMF_Summary" not in str(inputs[cell].value):
            issues.append(f"Inputs {cell} does not reference ExactPMF_Summary")

    # Board semantics: find headers by text instead of fixed cells
    headers = find_header_map(board, 1)
    required_headers = [
        "Market Home Line",
        "Fair Home Margin",
        "Home-Side Edge",
        "Market Total",
        "Fair Total",
        "Total Edge",
        "Home Win %",
        "Home Cover %",
        "Over %",
        "Fair ML Dec",
        "Fair ML US",
        "Fair Cover Dec",
        "Fair Cover US",
        "Fair Over Dec",
        "Fair Over US",
        "Fair Home Line",
    ]
    for h in required_headers:
        if h not in headers:
            issues.append(f"Board missing header '{h}'")

    # SpreadTotal exact PMF columns should be formula-driven in B/G
    if not is_formula(spread["B11"].value):
        issues.append("SpreadTotal!B11 should be formula-driven exact margin PMF")
    if not is_formula(spread["G11"].value):
        issues.append("SpreadTotal!G11 should be formula-driven exact total PMF")

    # Exact sheets should have data
    if summary.max_row < 2:
        issues.append("ExactPMF_Summary has no data rows")

    try:
        int(gp["B2"].value)
    except Exception:
        issues.append("GamePick_ModeA!B2 is not an integer selector")

    if issues:
        print("QA FAILED")
        for x in issues:
            print(" -", x)
        raise SystemExit(1)

    print("QA PASSED (EXACT MODE V3)")


if __name__ == "__main__":
    main()
