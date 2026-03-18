#!/usr/bin/env python3
from __future__ import annotations
import argparse
import shutil
from pathlib import Path
from openpyxl import load_workbook

def patch_inputs(ws):
    # identity block
    ws["B3"] = '=INDEX(GameInputs!$A:$A,GamePick_ModeA!$B$2+1)'
    ws["B4"] = '=INDEX(GameInputs!$B:$B,GamePick_ModeA!$B$2+1)'
    ws["B5"] = '=INDEX(GameInputs!$C:$C,GamePick_ModeA!$B$2+1)'
    ws["B6"] = '=INDEX(GameInputs!$D:$D,GamePick_ModeA!$B$2+1)'

    # league averages / controls
    ws["F3"] = '=AVERAGE(BlendedRatings!E2:E1000)'
    ws["F4"] = '=AVERAGE(BlendedRatings!F2:F1000)'
    ws["F5"] = '=AVERAGE(BlendedRatings!G2:G1000)'

    # team helper block
    ws["C7"]  = '=IFERROR(INDEX(BlendedRatings!$G:$G,MATCH($B$4,BlendedRatings!$A:$A,0)),"")'
    ws["C8"]  = '=IFERROR(INDEX(BlendedRatings!$E:$E,MATCH($B$4,BlendedRatings!$A:$A,0)),"")'
    ws["C9"]  = '=IFERROR(INDEX(BlendedRatings!$F:$F,MATCH($B$4,BlendedRatings!$A:$A,0)),"")'
    ws["C10"] = '=IFERROR(INDEX(BlendedRatings!$G:$G,MATCH($B$5,BlendedRatings!$A:$A,0)),"")'
    ws["C11"] = '=IFERROR(INDEX(BlendedRatings!$E:$E,MATCH($B$5,BlendedRatings!$A:$A,0)),"")'
    ws["C12"] = '=IFERROR(INDEX(BlendedRatings!$F:$F,MATCH($B$5,BlendedRatings!$A:$A,0)),"")'

    # rotation / market inputs from GameInputs
    ws["B10"] = '=IFERROR(INDEX(GameInputs!$G:$G,GamePick_ModeA!$B$2+1),0)'
    ws["B11"] = '=IFERROR(INDEX(GameInputs!$H:$H,GamePick_ModeA!$B$2+1),0)'
    ws["B18"] = '=IFERROR(INDEX(GameInputs!$E:$E,GamePick_ModeA!$B$2+1),"")'
    ws["B19"] = '=IFERROR(INDEX(GameInputs!$F:$F,GamePick_ModeA!$B$2+1),"")'

    # exact PMF summary values for the selected game row
    # ExactPMF_Summary layout:
    # A Cutoff, B Home Team, C Away Team, D Site, E Market Home Line, F Market Total,
    # G Expected Tempo, H Expected Home ORtg, I Expected Away ORtg,
    # J Expected Home Points, K Expected Away Points, L Fair Home Margin, M Fair Total,
    # N Exact Home Win Prob, O Exact Home Cover Prob, P Exact Over Prob,
    # Q/R/S decimal odds, T/U/V American odds.
    row_expr = 'GamePick_ModeA!$B$2+1'
    ws["B7"]  = f'=IFERROR(INDEX(ExactPMF_Summary!$G:$G,{row_expr}),"")'
    ws["B8"]  = f'=IFERROR(INDEX(ExactPMF_Summary!$H:$H,{row_expr}),"")'
    ws["B9"]  = f'=IFERROR(INDEX(ExactPMF_Summary!$I:$I,{row_expr}),"")'
    ws["B12"] = f'=IFERROR(INDEX(ExactPMF_Summary!$J:$J,{row_expr}),"")'
    ws["B13"] = f'=IFERROR(INDEX(ExactPMF_Summary!$K:$K,{row_expr}),"")'
    ws["B14"] = f'=IFERROR(INDEX(ExactPMF_Summary!$L:$L,{row_expr}),"")'
    ws["B15"] = f'=IFERROR(INDEX(ExactPMF_Summary!$M:$M,{row_expr}),"")'
    ws["B20"] = f'=IFERROR(INDEX(ExactPMF_Summary!$N:$N,{row_expr}),"")'
    ws["B21"] = f'=IFERROR(INDEX(ExactPMF_Summary!$O:$O,{row_expr}),"")'
    ws["B22"] = f'=IFERROR(INDEX(ExactPMF_Summary!$P:$P,{row_expr}),"")'
    ws["B25"] = f'=IFERROR(INDEX(ExactPMF_Summary!$Q:$Q,{row_expr}),"")'
    ws["B26"] = f'=IFERROR(INDEX(ExactPMF_Summary!$T:$T,{row_expr}),"")'
    ws["B27"] = f'=IFERROR(INDEX(ExactPMF_Summary!$R:$R,{row_expr}),"")'
    ws["B28"] = f'=IFERROR(INDEX(ExactPMF_Summary!$U:$U,{row_expr}),"")'
    ws["B29"] = f'=IFERROR(INDEX(ExactPMF_Summary!$S:$S,{row_expr}),"")'
    ws["B30"] = f'=IFERROR(INDEX(ExactPMF_Summary!$V:$V,{row_expr}),"")'

    # edges
    ws["B23"] = '=IF(AND(ISNUMBER(B14),ISNUMBER(B18)),B14+B18,"")'
    ws["B24"] = '=IF(AND(ISNUMBER(B15),ISNUMBER(B19)),B15-B19,"")'

def patch_dashboard(ws):
    ws["B2"] = '=Inputs!$B$5&" @ "&Inputs!$B$4'
    ws["E2"] = '=GamePick_ModeA!$B$2'

    ws["B4"] = '=Inputs!$B$18'
    ws["B5"] = '=Inputs!$B$14'
    ws["B6"] = '=Inputs!$B$23'

    ws["E4"] = '=Inputs!$B$19'
    ws["E5"] = '=Inputs!$B$15'
    ws["E6"] = '=Inputs!$B$24'

    ws["H4"] = '=Inputs!$B$20'
    ws["H5"] = '=Inputs!$B$21'
    ws["H6"] = '=Inputs!$B$22'

    ws["K4"] = '=Inputs!$B$26'
    ws["K5"] = '=Inputs!$B$28'
    ws["K6"] = '=Inputs!$B$30'

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True)
    args = ap.parse_args()

    p = Path(args.workbook)
    if not p.exists():
        raise SystemExit(f"Workbook not found: {p}")

    backup = p.with_suffix(p.suffix + ".prepatch.bak")
    if not backup.exists():
        shutil.copy2(p, backup)

    wb = load_workbook(p)

    need = ["Inputs", "MarketMaker_Dashboard", "ExactPMF_Summary", "GameInputs", "BlendedRatings", "GamePick_ModeA"]
    missing = [s for s in need if s not in wb.sheetnames]
    if missing:
        raise SystemExit(f"Workbook missing required sheet(s): {missing}")

    patch_inputs(wb["Inputs"])
    patch_dashboard(wb["MarketMaker_Dashboard"])

    wb.save(p)
    print(f"Patched workbook in place: {p}")
    print(f"Backup created at: {backup}")

if __name__ == "__main__":
    main()
