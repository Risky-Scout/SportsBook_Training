#!/usr/bin/env python3
from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


def clone_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy(src._style)
    dst.number_format = src.number_format
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)


def set_cell(ws, ref: str, value, template_ref: str | None = None, center: bool = False) -> None:
    ws[ref] = value
    if template_ref:
        clone_style(ws[template_ref], ws[ref])
    if center:
        ws[ref].alignment = Alignment(horizontal="center", vertical="center")


def fair_us_formula(prob_ref: str) -> str:
    # Converts probability -> fair American odds, blank if invalid.
    return (
        f'=IF(OR({prob_ref}="",{prob_ref}<=0,{prob_ref}>=1),"",'
        f'IF({prob_ref}>=0.5,ROUND(-100*{prob_ref}/(1-{prob_ref}),0),ROUND(100*(1-{prob_ref})/{prob_ref},0)))'
    )


def fair_dec_formula(prob_ref: str) -> str:
    return f'=IF(OR({prob_ref}="",{prob_ref}<=0,{prob_ref}>=1),"",1/{prob_ref})'


def main() -> None:
    ap = argparse.ArgumentParser(description="Clean SpreadTotal for exact-PMF trader-facing use.")
    ap.add_argument("--workbook", required=True)
    args = ap.parse_args()

    path = Path(args.workbook)
    wb = load_workbook(path)
    ws = wb["SpreadTotal"]

    # --- Top summary blocks ---
    set_cell(ws, "A3", "Home Team", "A3")
    set_cell(ws, "A4", "Away Team", "A4")
    set_cell(ws, "A5", "Market Home Line", "A5")
    set_cell(ws, "A6", "Fair Home Line", "A6")
    set_cell(ws, "A7", "Market Total", "A7")
    set_cell(ws, "A8", "Fair Total", "A8")

    ws["B3"] = "=Inputs!B4"
    ws["B4"] = "=Inputs!B5"
    ws["B5"] = "=Inputs!B18"
    ws["B6"] = "=-Inputs!B14"
    ws["B7"] = "=Inputs!B19"
    ws["B8"] = "=Inputs!B15"

    # Probability / edge block
    set_cell(ws, "C3", "Home Win %", "C3")
    set_cell(ws, "C4", "Home Cover %", "C4")
    set_cell(ws, "C5", "Over %", "C5")
    set_cell(ws, "C6", "Home-Side Edge", "C6")
    set_cell(ws, "C7", "Total Edge", "C7")
    set_cell(ws, "C8", "Confidence", "C8")

    ws["D3"] = "=Inputs!B20"
    ws["D4"] = "=Inputs!B21"
    ws["D5"] = "=Inputs!B22"
    ws["D6"] = "=Inputs!B23"
    ws["D7"] = "=Inputs!B24"
    ws["D8"] = '=IF(COUNT(D3:D7)=5,MIN(1,MAX(0,AVERAGE(ABS(D6)/6,ABS(D7)/10,D4,D5))),"")'

    # Fair odds block
    set_cell(ws, "F3", "Fair ML Dec", "F3")
    set_cell(ws, "F4", "Fair ML US", "F4")
    set_cell(ws, "F5", "Fair Cover Dec", "F5")
    set_cell(ws, "F6", "Fair Cover US", "F6")
    set_cell(ws, "F7", "Fair Over Dec", "F7")
    set_cell(ws, "F8", "Fair Over US", "F8")

    ws["G3"] = "=Inputs!B25"
    ws["G4"] = "=Inputs!B26"
    ws["G5"] = "=Inputs!B27"
    ws["G6"] = "=Inputs!B28"
    ws["G7"] = "=Inputs!B29"
    ws["G8"] = "=Inputs!B30"

    # Notes / sign convention
    set_cell(ws, "I3", "Interpretation", "I3")
    ws["I4"] = "Model uses exact PMF from ExactPMF sheets."
    ws["I5"] = "Margin chart x-axis = home margin."
    ws["I6"] = "Market marker on margin chart = - Market Home Line."
    ws["I7"] = "Model marker on margin chart = Fair Home Margin."
    ws["I8"] = "Visible trading line view is elsewhere in workbook."

    set_cell(ws, "N2", "Sign convention", "N2")
    ws["N3"] = "Fair Home Margin > 0 means home expected to win."
    ws["N4"] = "Fair Home Line = - Fair Home Margin."
    ws["N5"] = "Market Home Line: negative favorite / positive dog."

    # --- Margin PMF block ---
    set_cell(ws, "A10", "Home Margin", "A10", center=True)
    set_cell(ws, "B10", "Dist Prob", "B10", center=True)
    set_cell(ws, "C10", "Market Marker", "C10", center=True)
    set_cell(ws, "D10", "Model Marker", "D10", center=True)

    start_row = 11
    end_row = 61

    for r in range(start_row, end_row + 1):
        offset = r - start_row - 25
        ws[f"A{r}"] = offset
        ws[f"B{r}"] = (
            f'=IFERROR(SUMIFS(ExactPMF_Margin!$E:$E,'
            f'ExactPMF_Margin!$A:$A,Inputs!$B$3,'
            f'ExactPMF_Margin!$B:$B,Inputs!$B$4,'
            f'ExactPMF_Margin!$C:$C,Inputs!$B$5,'
            f'ExactPMF_Margin!$D:$D,A{r}),"")'
        )
        ws[f"C{r}"] = f'=IF(AND(ISNUMBER(B{r}),ABS(A{r}-ROUND(-Inputs!$B$18,0))<=0.5),B{r},"")'
        ws[f"D{r}"] = f'=IF(AND(ISNUMBER(B{r}),ABS(A{r}-ROUND(Inputs!$B$14,0))<=0.5),B{r},"")'

    # --- Total PMF block ---
    set_cell(ws, "F10", "Game Total", "F10", center=True)
    set_cell(ws, "G10", "Dist Prob", "G10", center=True)
    set_cell(ws, "H10", "Market Marker", "H10", center=True)
    set_cell(ws, "I10", "Model Marker", "I10", center=True)

    ws["F11"] = "=ROUND(MIN(Inputs!$B$15,Inputs!$B$19)-20,0)"
    for r in range(12, end_row + 1):
        ws[f"F{r}"] = f"=F{r-1}+1"
    for r in range(start_row, end_row + 1):
        ws[f"G{r}"] = (
            f'=IFERROR(SUMIFS(ExactPMF_Total!$E:$E,'
            f'ExactPMF_Total!$A:$A,Inputs!$B$3,'
            f'ExactPMF_Total!$B:$B,Inputs!$B$4,'
            f'ExactPMF_Total!$C:$C,Inputs!$B$5,'
            f'ExactPMF_Total!$D:$D,F{r}),"")'
        )
        ws[f"H{r}"] = f'=IF(AND(ISNUMBER(G{r}),ABS(F{r}-ROUND(Inputs!$B$19,0))<=0.5),G{r},"")'
        ws[f"I{r}"] = f'=IF(AND(ISNUMBER(G{r}),ABS(F{r}-ROUND(Inputs!$B$15,0))<=0.5),G{r},"")'

    # --- Alternate spread ladder ---
    set_cell(ws, "K10", "Alt Spread", "K10", center=True)
    set_cell(ws, "L10", "Home Cover %", "L10", center=True)
    set_cell(ws, "M10", "Fair Dec", "M10", center=True)
    set_cell(ws, "N10", "Fair US", "N10", center=True)

    ws["K11"] = "=ROUND(Inputs!$B$18,0)-6"
    for r in range(12, 26):
        ws[f"K{r}"] = f"=K{r-1}+1"
    for r in range(11, 26):
        ws[f"L{r}"] = (
            f'=IFERROR(SUMPRODUCT((ExactPMF_Margin!$A$2:$A$5000=Inputs!$B$3)*'
            f'(ExactPMF_Margin!$B$2:$B$5000=Inputs!$B$4)*'
            f'(ExactPMF_Margin!$C$2:$C$5000=Inputs!$B$5)*'
            f'(ExactPMF_Margin!$D$2:$D$5000>-K{r})*'
            f'ExactPMF_Margin!$E$2:$E$5000),"")'
        )
        ws[f"M{r}"] = fair_dec_formula(f"L{r}")
        ws[f"N{r}"] = fair_us_formula(f"L{r}")

    # --- Alternate total ladder ---
    set_cell(ws, "P10", "Alt Total", "P10", center=True)
    set_cell(ws, "Q10", "Over %", "Q10", center=True)
    set_cell(ws, "R10", "Fair Dec", "R10", center=True)
    set_cell(ws, "S10", "Fair US", "S10", center=True)

    ws["P11"] = "=ROUND(Inputs!$B$19,0)-7"
    for r in range(12, 26):
        ws[f"P{r}"] = f"=P{r-1}+1"
    for r in range(11, 26):
        ws[f"Q{r}"] = (
            f'=IFERROR(SUMPRODUCT((ExactPMF_Total!$A$2:$A$5000=Inputs!$B$3)*'
            f'(ExactPMF_Total!$B$2:$B$5000=Inputs!$B$4)*'
            f'(ExactPMF_Total!$C$2:$C$5000=Inputs!$B$5)*'
            f'(ExactPMF_Total!$D$2:$D$5000>P{r})*'
            f'ExactPMF_Total!$E$2:$E$5000),"")'
        )
        ws[f"R{r}"] = fair_dec_formula(f"Q{r}")
        ws[f"S{r}"] = fair_us_formula(f"Q{r}")

    # Formatting: percentages on probability columns
    for col in ["B", "C", "D", "G", "H", "I", "L", "Q", "D"]:
        for r in range(11, end_row + 1):
            ws[f"{col}{r}"].number_format = "0.0%"
    for cell in ["D3", "D4", "D5", "D8"]:
        ws[cell].number_format = "0.0%"
    for cell in ["L11:L25", "Q11:Q25"]:
        pass  # handled by row loops below
    for r in range(11, 26):
        ws[f"L{r}"].number_format = "0.0%"
        ws[f"Q{r}"].number_format = "0.0%"
    for r in range(11, end_row + 1):
        ws[f"B{r}"].number_format = "0.0%"
        ws[f"G{r}"].number_format = "0.0%"
        ws[f"C{r}"].number_format = "0.0%"
        ws[f"D{r}"].number_format = "0.0%"
        ws[f"H{r}"].number_format = "0.0%"
        ws[f"I{r}"].number_format = "0.0%"

    # Widths / readability
    widths = {
        "A": 10, "B": 9, "C": 11, "D": 11,
        "F": 10, "G": 9, "H": 11, "I": 11,
        "K": 10, "L": 11, "M": 9, "N": 9,
        "P": 10, "Q": 9, "R": 9, "S": 9
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(path)
    print(f"Patched SpreadTotal exact cleanup in {path}")


if __name__ == "__main__":
    main()
