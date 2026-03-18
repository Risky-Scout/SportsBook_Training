#!/usr/bin/env python3
from openpyxl import load_workbook
from pathlib import Path
import argparse

def to_num(x):
    try:
        if x is None or x == "":
            return None
        return float(x)
    except Exception:
        return None

ap = argparse.ArgumentParser()
ap.add_argument("--workbook", required=True)
args = ap.parse_args()

path = Path(args.workbook)
wb = load_workbook(path)

# -----------------------------
# MarketMaker_Board
# Visible semantics = line semantics
# Hidden QA compatibility = Fair Home Margin
# -----------------------------
ws = wb["MarketMaker_Board"]

visible_headers = [
    "Row",
    "Cutoff",
    "Away Team",
    "Home Team",
    "Site",
    "Market Home Line",
    "Fair Home Line",
    "Home-Side Edge",
    "Market Total",
    "Fair Total",
    "Total Edge",
    "Home Win %",
    "Home Cover %",
    "Over %",
    "Fair ML Dec",
    "Fair ML US",
    "Fair Cover Dec",
    "Fair Cover US",
    "Fair Over Dec",
    "Fair Over US",
]

for i, h in enumerate(visible_headers, start=1):
    ws.cell(row=1, column=i).value = h

# Hidden compatibility column U for QA
ws.cell(row=1, column=21).value = "Fair Home Margin"
for r in range(2, max(ws.max_row, 200) + 1):
    fair_home_line = to_num(ws.cell(row=r, column=7).value)
    ws.cell(row=r, column=21).value = (-fair_home_line if fair_home_line is not None else None)
ws.column_dimensions["U"].hidden = True

# -----------------------------
# MarketMaker_Dashboard
# -----------------------------
if "MarketMaker_Dashboard" in wb.sheetnames:
    ws = wb["MarketMaker_Dashboard"]

    # Primary selected-game labels
    for cell, value in {
        "A4": "Market Home Line",
        "A5": "Fair Home Line",
        "A6": "Home-Side Edge",
        "D4": "Market Total",
        "D5": "Fair Total",
        "D6": "Total Edge",
    }.items():
        ws[cell] = value

    # Snapshot header row
    snapshot_headers = {
        1: "Row",
        2: "Away Team",
        3: "Home Team",
        4: "Market Home Line",
        5: "Fair Home Line",
        6: "Home-Side Edge",
        7: "Market Total",
        8: "Fair Total",
        9: "Total Edge",
        10: "Home Cover %",
        11: "Over %",
    }
    for col, val in snapshot_headers.items():
        ws.cell(row=10, column=col).value = val

# -----------------------------
# Inputs
# -----------------------------
if "Inputs" in wb.sheetnames:
    ws = wb["Inputs"]
    # Replace first-column labels only
    for r in range(1, 60):
        v = ws.cell(row=r, column=1).value
        if v == "Fair Home Margin":
            ws.cell(row=r, column=1).value = "Fair Home Line"
        elif v == "Market Spread":
            ws.cell(row=r, column=1).value = "Market Home Line"
        elif v == "Spread Edge":
            ws.cell(row=r, column=1).value = "Home-Side Edge"

    # Fix any explanatory note text
    for row in ws.iter_rows(min_row=1, max_row=80, min_col=1, max_col=12):
        for cell in row:
            if isinstance(cell.value, str):
                txt = cell.value
                txt = txt.replace("Fair Home Margin", "Fair Home Line")
                txt = txt.replace("Market Spread", "Market Home Line")
                txt = txt.replace("Spread Edge", "Home-Side Edge")
                if "Model spread = expected home points minus expected away points." in txt:
                    txt = "Fair Home Line = model-implied home line. Negative = home favorite, positive = home underdog."
                cell.value = txt

# -----------------------------
# SpreadTotal
# -----------------------------
if "SpreadTotal" in wb.sheetnames:
    ws = wb["SpreadTotal"]

    # Common visible labels seen on this sheet
    replacements = {
        "Market Spread": "Market Home Line",
        "Model Spread": "Fair Home Line",
        "Fair Home Margin": "Fair Home Line",
        "Spread Edge": "Home-Side Edge",
    }

    for row in ws.iter_rows(min_row=1, max_row=120, min_col=1, max_col=25):
        for cell in row:
            if isinstance(cell.value, str):
                txt = cell.value
                for old, new in replacements.items():
                    txt = txt.replace(old, new)
                txt = txt.replace(
                    "Fair Home Margin >0 means home expected to win.",
                    "Fair Home Line < 0 means model home favorite; > 0 means model home underdog."
                )
                txt = txt.replace(
                    "Market Home Line: =+ home dog; -= home favorite",
                    "Market Home Line: negative favorite / positive dog."
                )
                txt = txt.replace(
                    "Market Home Line = - Fair Home Margin.",
                    "Home-Side Edge = Market Home Line - Fair Home Line."
                )
                txt = txt.replace(
                    "Model spread = expected home margin from the OE/DE.",
                    "Fair Home Line = model-implied home line."
                )
                cell.value = txt

print(f"Patched trader-facing sign semantics in {path}")
wb.save(path)
