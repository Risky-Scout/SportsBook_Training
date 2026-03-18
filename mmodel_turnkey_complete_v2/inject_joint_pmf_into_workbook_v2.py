
#!/usr/bin/env python3
from __future__ import annotations
import argparse, shutil
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

def write_df(ws, df: pd.DataFrame):
    ws.delete_rows(1, ws.max_row)
    headers = list(df.columns)
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c).value = h
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx).value = val

def get_or_create_sheet(wb, name: str):
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)
    return ws

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--base_workbook", required=True)
    ap.add_argument("--summary_csv", required=True)
    ap.add_argument("--margin_csv", required=True)
    ap.add_argument("--total_csv", required=True)
    ap.add_argument("--out_workbook", required=True)
    args = ap.parse_args()

    base = Path(args.base_workbook)
    out = Path(args.out_workbook)
    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(base, out)

    summary = pd.read_csv(args.summary_csv)
    margin = pd.read_csv(args.margin_csv)
    total = pd.read_csv(args.total_csv)

    wb = load_workbook(out)
    for name in ["ExactPMF_Summary", "ExactPMF_Margin", "ExactPMF_Total"]:
        if name in wb.sheetnames:
            ws = wb[name]
            wb.remove(ws)
    sum_ws = wb.create_sheet("ExactPMF_Summary")
    mar_ws = wb.create_sheet("ExactPMF_Margin")
    tot_ws = wb.create_sheet("ExactPMF_Total")

    write_df(sum_ws, summary)
    write_df(mar_ws, margin)
    write_df(tot_ws, total)

    # hide data sheets
    for name in ["ExactPMF_Summary", "ExactPMF_Margin", "ExactPMF_Total"]:
        wb[name].sheet_state = "hidden"

    # Inputs selected-game exact probabilities / fair odds
    inp = wb["Inputs"]
    inp["B20"] = '=IFERROR(INDEX(ExactPMF_Summary!$N:$N,GamePick_ModeA!$B$2+1),"")'
    inp["B21"] = '=IFERROR(INDEX(ExactPMF_Summary!$O:$O,GamePick_ModeA!$B$2+1),"")'
    inp["B22"] = '=IFERROR(INDEX(ExactPMF_Summary!$P:$P,GamePick_ModeA!$B$2+1),"")'
    inp["B25"] = '=IFERROR(INDEX(ExactPMF_Summary!$Q:$Q,GamePick_ModeA!$B$2+1),"")'
    inp["B26"] = '=IFERROR(INDEX(ExactPMF_Summary!$T:$T,GamePick_ModeA!$B$2+1),"")'
    inp["B27"] = '=IFERROR(INDEX(ExactPMF_Summary!$R:$R,GamePick_ModeA!$B$2+1),"")'
    inp["B28"] = '=IFERROR(INDEX(ExactPMF_Summary!$U:$U,GamePick_ModeA!$B$2+1),"")'
    inp["B29"] = '=IFERROR(INDEX(ExactPMF_Summary!$S:$S,GamePick_ModeA!$B$2+1),"")'
    inp["B30"] = '=IFERROR(INDEX(ExactPMF_Summary!$V:$V,GamePick_ModeA!$B$2+1),"")'

    # Board exact probabilities / odds and fair margin-total from exact summary
    if "MarketMaker_Board" in wb.sheetnames:
        board = wb["MarketMaker_Board"]
        max_rows = max(2, board.max_row)
        for r in range(2, max_rows + 1):
            board[f"G{r}"] = f'=IFERROR(INDEX(ExactPMF_Summary!$L:$L,$A{r}+1),"")'
            board[f"H{r}"] = f'=IF(AND(ISNUMBER(F{r}),ISNUMBER(G{r})),G{r}+F{r},"")'
            board[f"J{r}"] = f'=IFERROR(INDEX(ExactPMF_Summary!$M:$M,$A{r}+1),"")'
            board[f"K{r}"] = f'=IF(AND(ISNUMBER(I{r}),ISNUMBER(J{r})),J{r}-I{r},"")'
            board[f"L{r}"] = f'=IFERROR(INDEX(ExactPMF_Summary!$N:$N,$A{r}+1),"")'
            board[f"M{r}"] = f'=IFERROR(INDEX(ExactPMF_Summary!$O:$O,$A{r}+1),"")'
            board[f"N{r}"] = f'=IFERROR(INDEX(ExactPMF_Summary!$P:$P,$A{r}+1),"")'
            board[f"O{r}"] = f'=IFERROR(INDEX(ExactPMF_Summary!$Q:$Q,$A{r}+1),"")'
            board[f"P{r}"] = f'=IFERROR(INDEX(ExactPMF_Summary!$T:$T,$A{r}+1),"")'
            board[f"Q{r}"] = f'=IFERROR(INDEX(ExactPMF_Summary!$R:$R,$A{r}+1),"")'
            board[f"R{r}"] = f'=IFERROR(INDEX(ExactPMF_Summary!$U:$U,$A{r}+1),"")'
            board[f"S{r}"] = f'=IFERROR(INDEX(ExactPMF_Summary!$S:$S,$A{r}+1),"")'
            board[f"T{r}"] = f'=IFERROR(INDEX(ExactPMF_Summary!$V:$V,$A{r}+1),"")'


    # Dashboard selected-game block exact summary wiring
    if "MarketMaker_Dashboard" in wb.sheetnames:
        dash = wb["MarketMaker_Dashboard"]
        dash["B2"] = '=IFERROR(INDEX(ExactPMF_Summary!$C:$C,GamePick_ModeA!$B$2+1)&" @ "&INDEX(ExactPMF_Summary!$B:$B,GamePick_ModeA!$B$2+1),"")'
        dash["E2"] = '=GamePick_ModeA!$B$2'
        dash["B4"] = '=IFERROR(INDEX(ExactPMF_Summary!$E:$E,GamePick_ModeA!$B$2+1),"")'
        dash["B5"] = '=IFERROR(INDEX(ExactPMF_Summary!$L:$L,GamePick_ModeA!$B$2+1),"")'
        dash["B6"] = '=IF(AND(ISNUMBER(B4),ISNUMBER(B5)),B5+B4,"")'
        dash["E4"] = '=IFERROR(INDEX(ExactPMF_Summary!$F:$F,GamePick_ModeA!$B$2+1),"")'
        dash["E5"] = '=IFERROR(INDEX(ExactPMF_Summary!$M:$M,GamePick_ModeA!$B$2+1),"")'
        dash["E6"] = '=IF(AND(ISNUMBER(E4),ISNUMBER(E5)),E5-E4,"")'
        dash["H4"] = '=IFERROR(INDEX(ExactPMF_Summary!$N:$N,GamePick_ModeA!$B$2+1),"")'
        dash["H5"] = '=IFERROR(INDEX(ExactPMF_Summary!$O:$O,GamePick_ModeA!$B$2+1),"")'
        dash["H6"] = '=IFERROR(INDEX(ExactPMF_Summary!$P:$P,GamePick_ModeA!$B$2+1),"")'
        dash["K4"] = '=IFERROR(INDEX(ExactPMF_Summary!$T:$T,GamePick_ModeA!$B$2+1),"")'
        dash["K5"] = '=IFERROR(INDEX(ExactPMF_Summary!$U:$U,GamePick_ModeA!$B$2+1),"")'
        dash["K6"] = '=IFERROR(INDEX(ExactPMF_Summary!$V:$V,GamePick_ModeA!$B$2+1),"")'

    # SpreadTotal exact PMF lookups
    if "SpreadTotal" in wb.sheetnames:
        st = wb["SpreadTotal"]
        for r in range(11, st.max_row + 1):
            st[f"B{r}"] = (
                f'=SUMIFS(ExactPMF_Margin!$E:$E,'
                f'ExactPMF_Margin!$A:$A,Inputs!$B$3,'
                f'ExactPMF_Margin!$B:$B,Inputs!$B$4,'
                f'ExactPMF_Margin!$C:$C,Inputs!$B$5,'
                f'ExactPMF_Margin!$D:$D,A{r})'
            )
            st[f"G{r}"] = (
                f'=SUMIFS(ExactPMF_Total!$E:$E,'
                f'ExactPMF_Total!$A:$A,Inputs!$B$3,'
                f'ExactPMF_Total!$B:$B,Inputs!$B$4,'
                f'ExactPMF_Total!$C:$C,Inputs!$B$5,'
                f'ExactPMF_Total!$D:$D,F{r})'
            )
        st["J7"] = "Model PMF / win-cover-over use ExactPMF v2."

    wb.save(out)
    print(f"Wrote exact workbook: {out}")

if __name__ == "__main__":
    main()
