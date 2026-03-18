#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.shapes import GraphicalProperties


def build_margin_chart(ws, min_row: int, max_row: int):
    chart = ScatterChart()
    chart.title = "Home Margin Exact PMF | Market vs Model"
    chart.style = 2
    chart.width = 10.8
    chart.height = 5.4
    chart.x_axis.title = "Home margin (points)"
    chart.y_axis.title = "Probability mass"
    chart.y_axis.number_format = "0.0%"
    chart.legend.position = "r"

    xvalues = Reference(ws, min_col=27, min_row=min_row, max_row=max_row)   # AA
    dist_y = Reference(ws, min_col=28, min_row=min_row, max_row=max_row)    # AB
    market_y = Reference(ws, min_col=29, min_row=min_row, max_row=max_row)  # AC
    model_y = Reference(ws, min_col=30, min_row=min_row, max_row=max_row)   # AD

    dist = Series(dist_y, xvalues, title="Exact PMF")
    dist.marker.symbol = "none"
    dist.graphicalProperties = GraphicalProperties()
    dist.graphicalProperties.line.width = 22000
    chart.series.append(dist)

    market = Series(market_y, xvalues, title="Market")
    market.graphicalProperties = GraphicalProperties()
    market.graphicalProperties.line.noFill = True
    market.marker.symbol = "diamond"
    market.marker.size = 9
    chart.series.append(market)

    model = Series(model_y, xvalues, title="Model")
    model.graphicalProperties = GraphicalProperties()
    model.graphicalProperties.line.noFill = True
    model.marker.symbol = "triangle"
    model.marker.size = 10
    chart.series.append(model)

    return chart


def main():
    ap = argparse.ArgumentParser(description="Rebuild SpreadTotal margin chart using hidden helper columns.")
    ap.add_argument("--workbook", required=True)
    args = ap.parse_args()

    path = Path(args.workbook)
    wb = load_workbook(path)
    ws = wb["SpreadTotal"]

    # Do not touch visible ladders. Use hidden helper columns only.
    for col in ["AA", "AB", "AC", "AD"]:
        ws.column_dimensions[col].hidden = True

    ws["AA10"] = "Home Margin"
    ws["AB10"] = "Dist Prob"
    ws["AC10"] = "Market Marker"
    ws["AD10"] = "Model Marker"

    start_row, end_row = 11, 61

    # Center on market-implied home margin with about ±3 SD
    ws["AA11"] = "=ROUND(-Inputs!$B$18 - 3*Inputs!$B$16,0)"
    for r in range(12, end_row + 1):
        ws[f"AA{r}"] = f"=AA{r-1}+1"

    for r in range(start_row, end_row + 1):
        ws[f"AB{r}"] = (
            f'=IFERROR(SUMIFS(ExactPMF_Margin!$E:$E,'
            f'ExactPMF_Margin!$A:$A,Inputs!$B$3,'
            f'ExactPMF_Margin!$B:$B,Inputs!$B$4,'
            f'ExactPMF_Margin!$C:$C,Inputs!$B$5,'
            f'ExactPMF_Margin!$D:$D,AA{r}),"")'
        )
        ws[f"AC{r}"] = f'=IF(AND(ISNUMBER(AB{r}),AA{r}=ROUND(-Inputs!$B$18,0)),AB{r},"")'
        ws[f"AD{r}"] = f'=IF(AND(ISNUMBER(AB{r}),AA{r}=ROUND(Inputs!$B$14,0)),AB{r},"")'
        ws[f"AB{r}"].number_format = "0.0%"
        ws[f"AC{r}"].number_format = "0.0%"
        ws[f"AD{r}"].number_format = "0.0%"

    # Remove old charts and place one clean margin chart lower on the sheet.
    ws._charts = []
    chart = build_margin_chart(ws, start_row, end_row)
    ws.add_chart(chart, "J28")

    wb.save(path)
    print(f"Patched safe centered margin chart in {path}")


if __name__ == "__main__":
    main()
