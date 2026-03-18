#!/usr/bin/env python3
from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


def clone_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy(src._style)
    dst.number_format = src.number_format
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)


def set_header(ws, cell_ref: str, text: str, template_ref: str) -> None:
    ws[cell_ref] = text
    clone_style(ws[template_ref], ws[cell_ref])
    ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")


def main() -> None:
    ap = argparse.ArgumentParser(description="Patch workbook to use consistent home-line signs without circular refs.")
    ap.add_argument("--workbook", required=True)
    args = ap.parse_args()

    path = Path(args.workbook)
    wb = load_workbook(path)

    board = wb["MarketMaker_Board"]
    dash = wb["MarketMaker_Dashboard"]

    # Board layout:
    # F = Market Home Line (visible)
    # G = Fair Home Line   (visible, same sign convention as market home line)
    # H = Home-Side Edge   (visible)
    # I/J/K = total fields
    # T = Fair Home Margin (helper only)

    set_header(board, "F1", "Market Home Line", "F1")
    set_header(board, "G1", "Fair Home Line", "G1")
    set_header(board, "H1", "Home-Side Edge", "H1")
    set_header(board, "T1", "Fair Home Margin", "T1")

    # ExactPMF_Summary column L = Fair Home Margin.
    for r in range(2, board.max_row + 1):
        board[f"T{r}"] = f'=IFERROR(INDEX(ExactPMF_Summary!$L:$L,ROW()),"")'
        board[f"G{r}"] = f'=IFERROR(-T{r},"")'
        board[f"H{r}"] = f'=IF(AND(ISNUMBER(F{r}),ISNUMBER(G{r})),F{r}-G{r},"")'

    # Top selected-game block: line-vs-line
    dash["A4"] = "Market Home Line"
    dash["A5"] = "Fair Home Line"
    dash["A6"] = "Home-Side Edge"
    dash["B4"] = "=MarketMaker_Board!F2"
    dash["B5"] = "=MarketMaker_Board!G2"
    dash["B6"] = "=MarketMaker_Board!H2"

    # Snapshot table
    headers = {
        "A10": "Row",
        "B10": "Away Team",
        "C10": "Home Team",
        "D10": "Site",
        "E10": "Market Home Line",
        "F10": "Fair Home Line",
        "G10": "Home-Side Edge",
        "H10": "Market Total",
        "I10": "Fair Total",
        "J10": "Total Edge",
        "K10": "Home Cover %",
        "L10": "Over %",
        "M10": "Fair Home Margin",
    }
    for ref, text in headers.items():
        set_header(dash, ref, text, "E10")

    for r in range(11, 200):
        br = r - 9
        dash[f"A{r}"] = f"=MarketMaker_Board!A{br}"
        dash[f"B{r}"] = f"=MarketMaker_Board!C{br}"
        dash[f"C{r}"] = f"=MarketMaker_Board!D{br}"
        dash[f"D{r}"] = f"=MarketMaker_Board!E{br}"
        dash[f"E{r}"] = f"=MarketMaker_Board!F{br}"
        dash[f"F{r}"] = f"=MarketMaker_Board!G{br}"
        dash[f"G{r}"] = f"=MarketMaker_Board!H{br}"
        dash[f"H{r}"] = f"=MarketMaker_Board!I{br}"
        dash[f"I{r}"] = f"=MarketMaker_Board!J{br}"
        dash[f"J{r}"] = f"=MarketMaker_Board!K{br}"
        dash[f"K{r}"] = f"=MarketMaker_Board!M{br}"
        dash[f"L{r}"] = f"=MarketMaker_Board!N{br}"
        dash[f"M{r}"] = f"=MarketMaker_Board!T{br}"

    # Notes
    dash["M2"] = "Chart guide"
    dash["M3"] = "Distribution = model probability mass"
    dash["M4"] = "Market Marker = current market number"
    dash["M5"] = "Model Marker = model-implied number"
    dash["M6"] = "Home line: + = home dog, - = home favorite"
    dash["M7"] = "Fair Home Line uses the same sign convention as Market Home Line"
    dash["M8"] = "Fair Home Margin helper only: + home wins, - away wins"

    widths = {
        "A": 7, "B": 16, "C": 16, "D": 7,
        "E": 18, "F": 16, "G": 15,
        "H": 12, "I": 10, "J": 11,
        "K": 13, "L": 10, "M": 17,
    }
    for col, width in widths.items():
        dash.column_dimensions[col].width = width

    wb.save(path)
    print(f"Patched consistent home-line signs safely in {path}")


if __name__ == "__main__":
    main()
