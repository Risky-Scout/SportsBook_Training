#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.shapes import GraphicalProperties


def build_chart(ws, min_row: int, max_row: int):
    chart = ScatterChart()
    chart.title = "Home Margin Exact PMF | Centered on Market"
    chart.style = 2
    chart.width = 10.5
    chart.height = 5.2
    chart.x_axis.title = "Home margin (points)"
    chart.y_axis.title = "Probability mass"
    chart.y_axis.number_format = "0.0%"
    chart.legend.position = "r"

    xvalues = Reference(ws, min_col=1, min_row=min_row, max_row=max_row)   # A
    dist_y = Reference(ws, min_col=2, min_row=min_row, max_row=max_row)    # B
    market_y = Reference(ws, min_col=3, min_row=min_row, max_row=max_row)  # C
    model_y = Reference(ws, min_col=4, min_row=min_row, max_row=max_row)   # D

    dist = Series(dist_y, xvalues, title="Exact PMF")
    dist.marker.symbol = "none"
    dist.graphicalProperties = GraphicalProperties()
    dist.graphicalProperties.line.width = 22000
    chart.series.append(dist)

    market = Series(market_y, xvalues, title="Market")
    market.graphicalProperties = GraphicalProperties()
    market.graphicalProperties.line.noFill = True
    market.marker.symbol = "diamond"
    market.marker.size = 10
    chart.series.append(market)

    model = Series(model_y, xvalues, title="Model")
    model.graphicalProperties = GraphicalProperties()
    model.graphicalProperties.line.noFill = True
    model.marker.symbol = "triangle"
    model.marker.size = 11
    chart.series.append(model)

    return chart


def main():
    ap = argparse.ArgumentParser(description="Center SpreadTotal home-margin chart on market margin with ±3 SD window.")
    ap.add_argument("--workbook", required=True)
    args = ap.parse_args()

    path = Path(args.workbook)
    wb = load_workbook(path)
    ws = wb["SpreadTotal"]

    ws._charts = []

    ws["A10"] = "Home Margin"
    ws["B10"] = "Dist Prob"
    ws["C10"] = "Market Marker"
    ws["D10"] = "Model Marker"

    start_row, end_row = 11, 61

    ws["A11"] = "=ROUND(-Inputs!$B$18 - 3*Inputs!$B$16,0)"
    for r in range(12, end_row + 1):
        ws[f"A{r}"] = f"=A{r-1}+1"

    for r in range(start_row, end_row + 1):
        ws[f"B{r}"] = (
            f'=IFERROR(SUMIFS(ExactPMF_Margin!$E:$E,'
            f'ExactPMF_Margin!$A:$A,Inputs!$B$3,'
            f'ExactPMF_Margin!$B:$B,Inputs!$B$4,'
            f'ExactPMF_Margin!$C:$C,Inputs!$B$5,'
            f'ExactPMF_Margin!$D:$D,A{r}),"")'
        )
        ws[f"C{r}"] = f'=IF(AND(ISNUMBER(B{r}),A{r}=ROUND(-Inputs!$B$18,0)),B{r},"")'
        ws[f"D{r}"] = f'=IF(AND(ISNUMBER(B{r}),A{r}=ROUND(Inputs!$B$14,0)),B{r},"")'
        ws[f"B{r}"].number_format = "0.0%"
        ws[f"C{r}"].number_format = "0.0%"
        ws[f"D{r}"].number_format = "0.0%"

    ws["F10"] = "Chart meaning"
    ws["F11"] = "X-axis = home margin"
    ws["F12"] = "Y-axis = exact PMF probability mass"
    ws["F13"] = "Market marker sits at - Market Home Line"
    ws["F14"] = "Model marker sits at Fair Home Margin"
    ws["F15"] = "Window spans about ±3 margin SD around market"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["F"].width = 42

    chart = build_chart(ws, start_row, end_row)
    ws.add_chart(chart, "F18")

    wb.save(path)
    print(f"Patched centered market-vs-model home-margin chart in {path}")


if __name__ == "__main__":
    main()
