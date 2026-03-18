#!/usr/bin/env python3
from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


def clone_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy(src._style)
    dst.number_format = src.number_format
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)


def set_header(ws, cell_ref: str, text: str, template_ref: str) -> None:
    ws[cell_ref] = text
    clone_style(ws[template_ref], ws[cell_ref])
    ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")


def main() -> None:
    ap = argparse.ArgumentParser(description="Patch dashboard to trader-facing line-vs-line view.")
    ap.add_argument("--workbook", required=True)
    args = ap.parse_args()

    path = Path(args.workbook)
    wb = load_workbook(path)
    ws = wb["MarketMaker_Dashboard"]

    # Top selected-game block: line-vs-line, not line-vs-margin.
    ws["A4"] = "Market Home Line"
    ws["A5"] = "Fair Home Line"
    ws["A6"] = "Home-Side Edge"

    # Keep style by reusing existing cells.
    for ref in ["A4", "A5", "A6"]:
        ws[ref].alignment = Alignment(horizontal="left", vertical="center")

    # Selected game formulas should compare line vs line.
    # Market Home Line comes from board col F, Fair Home Line from board col T, edge from board col H
    ws["B4"] = "=MarketMaker_Board!F2"
    ws["B5"] = "=MarketMaker_Board!T2"
    ws["B6"] = "=MarketMaker_Board!H2"

    # Snapshot table: make main spread comparison line-vs-line
    # Keep helper "Fair Home Margin" visible at far right.
    header_template = "E10"
    headers = {
        "A10": "Row",
        "B10": "Away Team",
        "C10": "Home Team",
        "D10": "Site",
        "E10": "Market Home Line",
        "F10": "Fair Home Line",
        "G10": "Home-Side Edge",
        "H10": "Market Total",
        "I10": "Fair Total",
        "J10": "Total Edge",
        "K10": "Home Cover %",
        "L10": "Over %",
        "M10": "Fair Home Margin",
    }
    for ref, text in headers.items():
        set_header(ws, ref, text, header_template)

    # Wire rows from MarketMaker_Board.
    # Board headers:
    # A row, C away, D home, E site, F market home line, G fair home margin,
    # H home-side edge, I market total, J fair total, K total edge,
    # M home cover %, N over %, T fair home line.
    for r in range(11, 200):
        board_r = r - 9
        ws[f"A{r}"] = f"=MarketMaker_Board!A{board_r}"
        ws[f"B{r}"] = f"=MarketMaker_Board!C{board_r}"
        ws[f"C{r}"] = f"=MarketMaker_Board!D{board_r}"
        ws[f"D{r}"] = f"=MarketMaker_Board!E{board_r}"
        ws[f"E{r}"] = f"=MarketMaker_Board!F{board_r}"   # Market Home Line
        ws[f"F{r}"] = f"=MarketMaker_Board!T{board_r}"   # Fair Home Line
        ws[f"G{r}"] = f"=MarketMaker_Board!H{board_r}"   # Home-Side Edge
        ws[f"H{r}"] = f"=MarketMaker_Board!I{board_r}"   # Market Total
        ws[f"I{r}"] = f"=MarketMaker_Board!J{board_r}"   # Fair Total
        ws[f"J{r}"] = f"=MarketMaker_Board!K{board_r}"   # Total Edge
        ws[f"K{r}"] = f"=MarketMaker_Board!M{board_r}"   # Home Cover %
        ws[f"L{r}"] = f"=MarketMaker_Board!N{board_r}"   # Over %
        ws[f"M{r}"] = f"=MarketMaker_Board!G{board_r}"   # Fair Home Margin helper

    # Sign-convention notes: make them explicit and brief.
    ws["M2"] = "Chart guide"
    ws["M3"] = "Distribution = model probability mass"
    ws["M4"] = "Market Marker = current market number"
    ws["M5"] = "Model Marker = model-implied number"
    ws["M6"] = "Home Line: + = home dog, - = home favorite"
    ws["M7"] = "Fair Home Line compares directly to Market Home Line"
    ws["M8"] = "Fair Home Margin is helper only: + home wins, - away wins"

    # Widen columns for readability.
    widths = {
        "A": 7, "B": 16, "C": 16, "D": 7,
        "E": 18, "F": 16, "G": 15,
        "H": 12, "I": 10, "J": 11,
        "K": 13, "L": 10, "M": 17,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(path)
    print(f"Patched trader-facing dashboard line view in {path}")


if __name__ == "__main__":
    main()
