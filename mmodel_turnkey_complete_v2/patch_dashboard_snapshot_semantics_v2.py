#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from copy import copy


def copy_style(src, dst):
    if src.has_style:
        dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.font:
        dst.font = copy(src.font)
    if src.border:
        dst.border = copy(src.border)
    if src.protection:
        dst.protection = copy(src.protection)


def main():
    ap = argparse.ArgumentParser(description="Patch snapshot table semantics on exact workbook dashboard")
    ap.add_argument("--workbook", required=True)
    args = ap.parse_args()

    path = Path(args.workbook)
    wb = load_workbook(path)

    ws = wb["MarketMaker_Dashboard"]
    board = wb["MarketMaker_Board"]

    # Backup note is handled by caller workflow; patch workbook in place.

    # Existing snapshot table row 10 headers / data start row 11.
    # Make semantics explicit and add Fair Home Line helper column.
    headers = {
        "A10": "Row",
        "B10": "Away Team",
        "C10": "Home Team",
        "D10": "Site",
        "E10": "Market Home Line",
        "F10": "Fair Home Margin",
        "G10": "Home-Side Edge",
        "H10": "Market Total",
        "I10": "Fair Total",
        "J10": "Total Edge",
        "K10": "Home Cover %",
        "L10": "Over %",
        "M10": "Fair Home Line",
    }

    # Preserve style from prior header region where possible.
    template_header = ws["E10"]
    for cell, text in headers.items():
        ws[cell] = text
        copy_style(template_header, ws[cell])
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")

    # Wire dashboard snapshot rows directly from MarketMaker_Board
    # Board row i maps to dashboard row i+9 (board data starts row 2).
    for r in range(11, 200):
        i = r - 9
        ws[f"A{r}"] = f"=MarketMaker_Board!A{i}"
        ws[f"B{r}"] = f"=MarketMaker_Board!C{i}"
        ws[f"C{r}"] = f"=MarketMaker_Board!D{i}"
        ws[f"D{r}"] = f"=MarketMaker_Board!E{i}"
        ws[f"E{r}"] = f"=MarketMaker_Board!F{i}"
        ws[f"F{r}"] = f"=MarketMaker_Board!G{i}"
        ws[f"G{r}"] = f"=MarketMaker_Board!H{i}"
        ws[f"H{r}"] = f"=MarketMaker_Board!I{i}"
        ws[f"I{r}"] = f"=MarketMaker_Board!J{i}"
        ws[f"J{r}"] = f"=MarketMaker_Board!K{i}"
        ws[f"K{r}"] = f"=MarketMaker_Board!M{i}"
        ws[f"L{r}"] = f"=MarketMaker_Board!N{i}"
        ws[f"M{r}"] = f"=MarketMaker_Board!T{i}"

    # Add a direct note clarifying sign convention
    ws["M2"] = "Chart guide"
    ws["M3"] = "Distribution = model probability mass"
    ws["M4"] = "Market Marker = current market number"
    ws["M5"] = "Model Marker = model-implied number"
    ws["M6"] = "Market Home Line: + = home dog, - = home favorite"
    ws["M7"] = "Fair Home Margin: + = home wins, - = away wins"
    ws["M8"] = "Fair Home Line = - Fair Home Margin"

    # Widen columns a bit for readability
    widths = {
        "B": 16, "C": 16, "D": 8, "E": 18, "F": 18, "G": 15,
        "H": 12, "I": 12, "J": 11, "K": 13, "L": 10, "M": 15
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(path)
    print(f"Patched dashboard snapshot semantics in {path}")


if __name__ == "__main__":
    main()
