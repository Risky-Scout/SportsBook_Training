#!/usr/bin/env python3
from openpyxl import load_workbook
import argparse
from pathlib import Path

EXPECTED = [
    "Row",
    "Cutoff",
    "Away Team",
    "Home Team",
    "Site",
    "Market Home Line",
    "Fair Home Line",
    "Home-Side Edge",
    "Market Total",
    "Fair Total",
    "Total Edge",
    "Home Win %",
    "Home Cover %",
    "Over %",
    "Fair ML Dec",
    "Fair ML US",
    "Fair Cover Dec",
    "Fair Cover US",
    "Fair Over Dec",
    "Fair Over US",
]

ap = argparse.ArgumentParser()
ap.add_argument("--workbook", required=True)
args = ap.parse_args()

path = Path(args.workbook)
wb = load_workbook(path)
ws = wb["MarketMaker_Board"]

for i, header in enumerate(EXPECTED, start=1):
    ws.cell(row=1, column=i).value = header

wb.save(path)
print(f"Patched MarketMaker_Board headers in {path}")
