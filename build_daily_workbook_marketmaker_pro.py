#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

TITLE_FILL = PatternFill("solid", fgColor="17365D")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
KPI_FILL = PatternFill("solid", fgColor="E2F0D9")
NOTE_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def hdr(cell, title=False):
    cell.fill = TITLE_FILL if title else SECTION_FILL
    cell.font = Font(color="FFFFFF", bold=True) if title else Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.border = BORDER


def write_df(ws, df: pd.DataFrame):
    for j, col in enumerate(df.columns, 1):
        c = ws.cell(1, j, str(col))
        hdr(c, title=True)
        ws.column_dimensions[get_column_letter(j)].width = 16
    for i, row in enumerate(df.itertuples(index=False), 2):
        for j, val in enumerate(row, 1):
            ws.cell(i, j, None if pd.isna(val) else val)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def add_sheet(wb, name):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    return ws


def build_selector(ws, n_games):
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22

    ws.merge_cells("A1:B1")
    ws["A1"] = "Game Selector"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    rows = [
        ("A2", "Selected row #", "B2", 1),
        ("A4", "Cutoff", "B4", '=INDEX(GameInputs!$A:$A,$B$2+1)'),
        ("A5", "Home Team", "B5", '=INDEX(GameInputs!$B:$B,$B$2+1)'),
        ("A6", "Away Team", "B6", '=INDEX(GameInputs!$C:$C,$B$2+1)'),
        ("A7", "Site", "B7", '=INDEX(GameInputs!$D:$D,$B$2+1)'),
        ("A8", "Market Spread", "B8", '=INDEX(GameInputs!$E:$E,$B$2+1)'),
        ("A9", "Market Total", "B9", '=INDEX(GameInputs!$F:$F,$B$2+1)'),
        ("A10", "Home RotΔ", "B10", '=INDEX(GameInputs!$G:$G,$B$2+1)'),
        ("A11", "Away RotΔ", "B11", '=INDEX(GameInputs!$H:$H,$B$2+1)'),
        ("A13", "Valid rows", "B13", f"1 to {n_games}"),
    ]
    for lc, ll, vc, val in rows:
        ws[lc] = ll
        ws[lc].fill = INPUT_FILL if lc == "A2" else SECTION_FILL
        ws[lc].font = Font(bold=True)
        ws[lc].border = BORDER
        ws[vc] = val
        ws[vc].fill = INPUT_FILL if vc == "B2" else KPI_FILL
        ws[vc].border = BORDER

    ws["A15"] = "Usage"
    ws["A15"].fill = TITLE_FILL
    ws["A15"].font = Font(color="FFFFFF", bold=True)
    ws["A16"] = "Change B2 to switch the selected game."
    ws["A17"] = "All dashboard and pricing tables update off that row."
    ws.freeze_panes = "A2"


def build_inputs(ws):
    for col, width in {"A":28, "B":16, "C":14, "E":26, "F":14}.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells("A1:F1")
    ws["A1"] = "Model Input Engine"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    labels = [
        (3, "Cutoff"),
        (4, "Home Team"),
        (5, "Away Team"),
        (6, "Site"),
        (7, "Expected Tempo"),
        (8, "Expected Home ORtg"),
        (9, "Expected Away ORtg"),
        (10, "Home RotΔ"),
        (11, "Away RotΔ"),
        (12, "Expected Home Pts"),
        (13, "Expected Away Pts"),
        (14, "Model Spread (home margin)"),
        (15, "Model Total"),
        (16, "Margin SD"),
        (17, "Total SD"),
        (18, "Market Spread"),
        (19, "Market Total"),
        (20, "Home Win %"),
        (21, "Home Cover % @ market spread"),
        (22, "Over % @ market total"),
        (23, "Spread Edge"),
        (24, "Total Edge"),
        (25, "Fair ML Decimal"),
        (26, "Fair ML American"),
        (27, "Fair Cover Decimal"),
        (28, "Fair Cover American"),
        (29, "Fair Over Decimal"),
        (30, "Fair Over American"),
    ]
    for r, label in labels:
        ws[f"A{r}"] = label
        ws[f"A{r}"].fill = SECTION_FILL
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"A{r}"].border = BORDER
        ws[f"B{r}"].border = BORDER

    refs = {
        "B3": "=GamePick_ModeA!$B$4",
        "B4": "=GamePick_ModeA!$B$5",
        "B5": "=GamePick_ModeA!$B$6",
        "B6": "=GamePick_ModeA!$B$7",
        "B10": "=GamePick_ModeA!$B$10",
        "B11": "=GamePick_ModeA!$B$11",
        "B18": "=GamePick_ModeA!$B$8",
        "B19": "=GamePick_ModeA!$B$9",
        "C7": '=IFERROR(INDEX(BlendedRatings!$G:$G,MATCH($B$4,BlendedRatings!$A:$A,0)),"")',
        "C8": '=IFERROR(INDEX(BlendedRatings!$E:$E,MATCH($B$4,BlendedRatings!$A:$A,0)),"")',
        "C9": '=IFERROR(INDEX(BlendedRatings!$F:$F,MATCH($B$4,BlendedRatings!$A:$A,0)),"")',
        "C10": '=IFERROR(INDEX(BlendedRatings!$G:$G,MATCH($B$5,BlendedRatings!$A:$A,0)),"")',
        "C11": '=IFERROR(INDEX(BlendedRatings!$E:$E,MATCH($B$5,BlendedRatings!$A:$A,0)),"")',
        "C12": '=IFERROR(INDEX(BlendedRatings!$F:$F,MATCH($B$5,BlendedRatings!$A:$A,0)),"")',
    }
    for cell, formula in refs.items():
        ws[cell] = formula

    cfg = [
        ("E3", "League Avg OE", '=AVERAGE(BlendedRatings!E2:E1000)'),
        ("E4", "League Avg DE", '=AVERAGE(BlendedRatings!F2:F1000)'),
        ("E5", "League Avg Tempo", '=AVERAGE(BlendedRatings!G2:G1000)'),
        ("E6", "Home-court adj (pts/100)", 1.5),
        ("E7", "RotΔ weight", 0.25),
        ("E8", "Margin SD base", 11.0),
        ("E9", "Margin SD rot add", 0.20),
        ("E10", "Total SD base", 14.5),
        ("E11", "Total SD pace add", 0.25),
    ]
    for cell, label, value in cfg:
        ws[cell] = label
        ws[cell].fill = SECTION_FILL
        ws[cell].font = Font(bold=True)
        ws[cell].border = BORDER
        out = "F" + cell[1:]
        ws[out] = value
        ws[out].fill = KPI_FILL if isinstance(value, str) and value.startswith("=") else INPUT_FILL
        ws[out].border = BORDER

    site = 'IF(OR($B$6="H",$B$6="Home",$B$6="home"),$F$6,IF(OR($B$6="A",$B$6="Away",$B$6="away"),-$F$6,0))'
    ws["B7"] = '=IF(AND(ISNUMBER(C7),ISNUMBER(C10),C7>0,C10>0,ISNUMBER($F$5)),0.85*(2/(1/C7+1/C10))+0.15*$F$5+IFERROR(INDEX(GameInputs!$K:$K,GamePick_ModeA!$B$2+1),0)+IFERROR(INDEX(GameInputs!$L:$L,GamePick_ModeA!$B$2+1),0),"")'
    ws["B8"] = f'=IF(AND(ISNUMBER(C8),ISNUMBER(C12),ISNUMBER($F$3),ISNUMBER($F$4)),$F$3+0.55*(C8-$F$3)+0.45*(C12-$F$4)+{site}+$F$7*($B$10-$B$11)+IFERROR(INDEX(GameInputs!$I:$I,GamePick_ModeA!$B$2+1),0),"")'
    ws["B9"] = f'=IF(AND(ISNUMBER(C11),ISNUMBER(C9),ISNUMBER($F$3),ISNUMBER($F$4)),$F$3+0.55*(C11-$F$3)+0.45*(C9-$F$4)-{site}+$F$7*($B$11-$B$10)+IFERROR(INDEX(GameInputs!$J:$J,GamePick_ModeA!$B$2+1),0),"")'
    ws["B12"] = '=IF(AND(ISNUMBER($B$7),ISNUMBER($B$8)),$B$7*$B$8/100,"")'
    ws["B13"] = '=IF(AND(ISNUMBER($B$7),ISNUMBER($B$9)),$B$7*$B$9/100,"")'
    ws["B14"] = '=IF(AND(ISNUMBER($B$12),ISNUMBER($B$13)),$B$12-$B$13,"")'
    ws["B15"] = '=IF(AND(ISNUMBER($B$12),ISNUMBER($B$13)),$B$12+$B$13,"")'
    ws["B16"] = '=IF(ISNUMBER($B$14),$F$8+$F$9*ABS($B$10-$B$11),"")'
    ws["B17"] = '=IF(ISNUMBER($B$15),$F$10+$F$11*ABS($B$7-$F$5),"")'
    ws["B20"] = '=IF(AND(ISNUMBER($B$14),ISNUMBER($B$16),$B$16>0),1-NORMDIST(0,$B$14,$B$16,TRUE),"")'
    ws["B21"] = '=IF(AND(ISNUMBER($B$18),ISNUMBER($B$14),ISNUMBER($B$16),$B$16>0),1-NORMDIST(-$B$18,$B$14,$B$16,TRUE),"")'
    ws["B22"] = '=IF(AND(ISNUMBER($B$19),ISNUMBER($B$15),ISNUMBER($B$17),$B$17>0),1-NORMDIST($B$19,$B$15,$B$17,TRUE),"")'
    ws["B23"] = '=IF(AND(ISNUMBER($B$14),ISNUMBER($B$18)),$B$14-$B$18,"")'
    ws["B24"] = '=IF(AND(ISNUMBER($B$15),ISNUMBER($B$19)),$B$15-$B$19,"")'
    ws["B25"] = '=IF(AND(ISNUMBER($B$20),$B$20>0,$B$20<1),1/$B$20,"")'
    ws["B26"] = '=IF(AND(ISNUMBER($B$20),$B$20>0,$B$20<1),IF($B$20>=0.5,-100*$B$20/(1-$B$20),100*(1-$B$20)/$B$20),"")'
    ws["B27"] = '=IF(AND(ISNUMBER($B$21),$B$21>0,$B$21<1),1/$B$21,"")'
    ws["B28"] = '=IF(AND(ISNUMBER($B$21),$B$21>0,$B$21<1),IF($B$21>=0.5,-100*$B$21/(1-$B$21),100*(1-$B$21)/$B$21),"")'
    ws["B29"] = '=IF(AND(ISNUMBER($B$22),$B$22>0,$B$22<1),1/$B$22,"")'
    ws["B30"] = '=IF(AND(ISNUMBER($B$22),$B$22>0,$B$22<1),IF($B$22>=0.5,-100*$B$22/(1-$B$22),100*(1-$B$22)/$B$22),"")'

    for r in range(20, 31):
        ws[f"B{r}"].fill = KPI_FILL
        ws[f"B{r}"].font = Font(bold=True)

    ws["D13"] = "Model spread = expected home points minus expected away points."
    ws["D14"] = "Use it as a model number, not as a claimed market-close truth."
    ws["D15"] = "Probabilities and fair odds are generated off model spread/total and SD assumptions."
    ws.freeze_panes = "A3"


def build_board(ws, n_games):
    headers = [
        "Row", "Cutoff", "Away Team", "Home Team", "Site",
        "Mkt Spr", "Model Spr", "Spr Edge",
        "Mkt Tot", "Model Tot", "Tot Edge",
        "Home Win %", "Home Cover %", "Over %",
        "Fair ML Dec", "Fair ML US",
        "Cover Dec", "Cover US",
        "Over Dec", "Over US",
        "Confidence"
    ]
    widths = [8, 12, 24, 24, 8, 12, 12, 11, 12, 12, 11, 12, 13, 10, 12, 12, 12, 12, 12, 12, 12]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        ws.cell(1, c, h)
        hdr(ws.cell(1, c), title=True)
        ws.column_dimensions[get_column_letter(c)].width = w

    hidden_headers = ["hOE", "hDE", "hT", "aOE", "aDE", "aT", "pace", "hExpO", "aExpO", "mSD", "tSD"]
    for c, h in enumerate(hidden_headers, 22):
        ws.cell(1, c, h)
        ws.column_dimensions[get_column_letter(c)].hidden = True

    for r in range(2, n_games + 2):
        site = f'IF(OR($E{r}="H",$E{r}="Home",$E{r}="home"),Inputs!$F$6,IF(OR($E{r}="A",$E{r}="Away",$E{r}="away"),-Inputs!$F$6,0))'
        formulas = {
            f"A{r}": r - 1,
            f"B{r}": f"=INDEX(GameInputs!$A:$A,$A{r}+1)",
            f"C{r}": f"=INDEX(GameInputs!$C:$C,$A{r}+1)",
            f"D{r}": f"=INDEX(GameInputs!$B:$B,$A{r}+1)",
            f"E{r}": f"=INDEX(GameInputs!$D:$D,$A{r}+1)",
            f"F{r}": f"=INDEX(GameInputs!$E:$E,$A{r}+1)",
            f"I{r}": f"=INDEX(GameInputs!$F:$F,$A{r}+1)",

            f"V{r}": f'=IFERROR(INDEX(BlendedRatings!$E:$E,MATCH($D{r},BlendedRatings!$A:$A,0)),"")',
            f"W{r}": f'=IFERROR(INDEX(BlendedRatings!$F:$F,MATCH($D{r},BlendedRatings!$A:$A,0)),"")',
            f"X{r}": f'=IFERROR(INDEX(BlendedRatings!$G:$G,MATCH($D{r},BlendedRatings!$A:$A,0)),"")',
            f"Y{r}": f'=IFERROR(INDEX(BlendedRatings!$E:$E,MATCH($C{r},BlendedRatings!$A:$A,0)),"")',
            f"Z{r}": f'=IFERROR(INDEX(BlendedRatings!$F:$F,MATCH($C{r},BlendedRatings!$A:$A,0)),"")',
            f"AA{r}": f'=IFERROR(INDEX(BlendedRatings!$G:$G,MATCH($C{r},BlendedRatings!$A:$A,0)),"")',

            f"AB{r}": f'=IF(AND(ISNUMBER(X{r}),ISNUMBER(AA{r}),X{r}>0,AA{r}>0,ISNUMBER(Inputs!$F$5)),0.85*(2/(1/X{r}+1/AA{r}))+0.15*Inputs!$F$5+INDEX(GameInputs!$K:$K,$A{r}+1)+INDEX(GameInputs!$L:$L,$A{r}+1),"")',
            f"AC{r}": f'=IF(AND(ISNUMBER(V{r}),ISNUMBER(Z{r}),ISNUMBER(Inputs!$F$3),ISNUMBER(Inputs!$F$4)),Inputs!$F$3+0.55*(V{r}-Inputs!$F$3)+0.45*(Z{r}-Inputs!$F$4)+{site}+Inputs!$F$7*(INDEX(GameInputs!$G:$G,$A{r}+1)-INDEX(GameInputs!$H:$H,$A{r}+1))+INDEX(GameInputs!$I:$I,$A{r}+1),"")',
            f"AD{r}": f'=IF(AND(ISNUMBER(Y{r}),ISNUMBER(W{r}),ISNUMBER(Inputs!$F$3),ISNUMBER(Inputs!$F$4)),Inputs!$F$3+0.55*(Y{r}-Inputs!$F$3)+0.45*(W{r}-Inputs!$F$4)-{site}+Inputs!$F$7*(INDEX(GameInputs!$H:$H,$A{r}+1)-INDEX(GameInputs!$G:$G,$A{r}+1))+INDEX(GameInputs!$J:$J,$A{r}+1),"")',
            f"AE{r}": f'=IF(ISNUMBER(AB{r}),Inputs!$F$8+Inputs!$F$9*ABS(INDEX(GameInputs!$G:$G,$A{r}+1)-INDEX(GameInputs!$H:$H,$A{r}+1)),"")',
            f"AF{r}": f'=IF(ISNUMBER(AB{r}),Inputs!$F$10+Inputs!$F$11*ABS(AB{r}-Inputs!$F$5),"")',

            f"G{r}": f'=IF(AND(ISNUMBER(AB{r}),ISNUMBER(AC{r}),ISNUMBER(AD{r})),AB{r}*(AC{r}-AD{r})/100,"")',
            f"H{r}": f'=IF(AND(ISNUMBER(G{r}),ISNUMBER(F{r})),G{r}-F{r},"")',
            f"J{r}": f'=IF(AND(ISNUMBER(AB{r}),ISNUMBER(AC{r}),ISNUMBER(AD{r})),AB{r}*(AC{r}+AD{r})/100,"")',
            f"K{r}": f'=IF(AND(ISNUMBER(J{r}),ISNUMBER(I{r})),J{r}-I{r},"")',
            f"L{r}": f'=IF(AND(ISNUMBER(G{r}),ISNUMBER(AE{r}),AE{r}>0),1-NORMDIST(0,G{r},AE{r},TRUE),"")',
            f"M{r}": f'=IF(AND(ISNUMBER(F{r}),ISNUMBER(G{r}),ISNUMBER(AE{r}),AE{r}>0),1-NORMDIST(-F{r},G{r},AE{r},TRUE),"")',
            f"N{r}": f'=IF(AND(ISNUMBER(I{r}),ISNUMBER(J{r}),ISNUMBER(AF{r}),AF{r}>0),1-NORMDIST(I{r},J{r},AF{r},TRUE),"")',
            f"O{r}": f'=IF(AND(ISNUMBER(L{r}),L{r}>0,L{r}<1),1/L{r},"")',
            f"P{r}": f'=IF(AND(ISNUMBER(L{r}),L{r}>0,L{r}<1),IF(L{r}>=0.5,-100*L{r}/(1-L{r}),100*(1-L{r})/L{r}),"")',
            f"Q{r}": f'=IF(AND(ISNUMBER(M{r}),M{r}>0,M{r}<1),1/M{r},"")',
            f"R{r}": f'=IF(AND(ISNUMBER(M{r}),M{r}>0,M{r}<1),IF(M{r}>=0.5,-100*M{r}/(1-M{r}),100*(1-M{r})/M{r}),"")',
            f"S{r}": f'=IF(AND(ISNUMBER(N{r}),N{r}>0,N{r}<1),1/N{r},"")',
            f"T{r}": f'=IF(AND(ISNUMBER(N{r}),N{r}>0,N{r}<1),IF(N{r}>=0.5,-100*N{r}/(1-N{r}),100*(1-N{r})/N{r}),"")',
            f"U{r}": f'=IF(AND(ISNUMBER(H{r}),ISNUMBER(K{r})),MIN(1,(ABS(H{r})/3)+(ABS(K{r})/8)),"")',
        }
        for cell, formula in formulas.items():
            ws[cell] = formula

        for c in range(1, 22):
            ws.cell(r, c).border = BORDER

        for c in "FGHIJK":
            ws[f"{c}{r}"].number_format = "0.0"
        for c in "LMN":
            ws[f"{c}{r}"].number_format = "0.0%"
        ws[f"O{r}"].number_format = "0.00"
        ws[f"Q{r}"].number_format = "0.00"
        ws[f"S{r}"].number_format = "0.00"
        ws[f"P{r}"].number_format = "0"
        ws[f"R{r}"].number_format = "0"
        ws[f"T{r}"].number_format = "0"
        ws[f"U{r}"].number_format = "0.0%"

    ws.freeze_panes = "A2"
    ws.conditional_formatting.add(
        f"H2:H{n_games+1}",
        ColorScaleRule(start_type="min", start_color="F8696B", mid_type="num", mid_value=0, mid_color="FFEB84", end_type="max", end_color="63BE7B")
    )
    ws.conditional_formatting.add(
        f"K2:K{n_games+1}",
        ColorScaleRule(start_type="min", start_color="F8696B", mid_type="num", mid_value=0, mid_color="FFEB84", end_type="max", end_color="63BE7B")
    )


def build_spreadtotal(ws):
    widths = {
        "A": 12, "B": 12, "C": 12, "D": 12,
        "F": 12, "G": 12, "H": 12, "I": 12,
        "K": 12, "L": 12, "M": 12, "N": 12,
        "P": 12, "Q": 12, "R": 12, "S": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells("A1:S1")
    ws["A1"] = "Distribution, Pricing, and Fair Odds"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    summary = [
        ("A3", "Home Team", "B3", '=Inputs!$B$4', "@"),
        ("A4", "Away Team", "B4", '=Inputs!$B$5', "@"),
        ("A5", "Market Spread", "B5", '=Inputs!$B$18', "0.0"),
        ("A6", "Model Spread", "B6", '=Inputs!$B$14', "0.0"),
        ("A7", "Market Total", "B7", '=Inputs!$B$19', "0.0"),
        ("A8", "Model Total", "B8", '=Inputs!$B$15', "0.0"),

        ("D3", "Home Win %", "E3", '=Inputs!$B$20', "0.0%"),
        ("D4", "Home Cover %", "E4", '=Inputs!$B$21', "0.0%"),
        ("D5", "Over %", "E5", '=Inputs!$B$22', "0.0%"),
        ("D6", "Spread Edge", "E6", '=Inputs!$B$23', "0.0"),
        ("D7", "Total Edge", "E7", '=Inputs!$B$24', "0.0"),
        ("D8", "Confidence", "E8", '=MIN(1,(ABS(Inputs!$B$23)/3)+(ABS(Inputs!$B$24)/8))', "0.0%"),

        ("G3", "Fair ML Dec", "H3", '=Inputs!$B$25', "0.00"),
        ("G4", "Fair ML US", "H4", '=Inputs!$B$26', "0"),
        ("G5", "Fair Cover Dec", "H5", '=Inputs!$B$27', "0.00"),
        ("G6", "Fair Cover US", "H6", '=Inputs!$B$28', "0"),
        ("G7", "Fair Over Dec", "H7", '=Inputs!$B$29', "0.00"),
        ("G8", "Fair Over US", "H8", '=Inputs!$B$30', "0"),
    ]
    for lc, ll, vc, fo, nf in summary:
        ws[lc] = ll
        ws[lc].fill = SECTION_FILL
        ws[lc].font = Font(bold=True)
        ws[lc].border = BORDER
        ws[vc] = fo
        ws[vc].fill = KPI_FILL
        ws[vc].font = Font(bold=True)
        ws[vc].number_format = nf
        ws[vc].border = BORDER

    ws["J3"] = "Interpretation"
    ws["J3"].fill = TITLE_FILL
    ws["J3"].font = Font(color="FFFFFF", bold=True)
    ws["J4"] = "Model spread = expected home margin from the OE/DE + tempo model."
    ws["J5"] = "Market spread / total come from Schedule.csv."
    ws["J6"] = "Fair odds are no-vig probabilities off the model, not sportsbook prices."
    ws["J7"] = "Charts show distribution + market marker + model marker."
    for c in ["J4", "J5", "J6", "J7"]:
        ws[c].fill = NOTE_FILL
        ws[c].border = BORDER

    headers = {
        "A10": "Home Margin",
        "B10": "Dist Prob",
        "C10": "Market Marker",
        "D10": "Model Marker",
        "F10": "Game Total",
        "G10": "Dist Prob",
        "H10": "Market Marker",
        "I10": "Model Marker",
        "K10": "Alt Spread",
        "L10": "Home Cover %",
        "M10": "Fair Dec",
        "N10": "Fair US",
        "P10": "Alt Total",
        "Q10": "Over %",
        "R10": "Fair Dec",
        "S10": "Fair US",
    }
    for cell, label in headers.items():
        ws[cell] = label
        hdr(ws[cell], title=True)

    for r in range(11, 43):
        idx = r - 11
        ws[f"A{r}"] = f'=ROUND(Inputs!$B$14,0)-15+{idx}'
        ws[f"B{r}"] = f'=IF(AND(ISNUMBER(A{r}),ISNUMBER(Inputs!$B$14),ISNUMBER(Inputs!$B$16),Inputs!$B$16>0),NORMDIST(A{r}+0.5,Inputs!$B$14,Inputs!$B$16,TRUE)-NORMDIST(A{r}-0.5,Inputs!$B$14,Inputs!$B$16,TRUE),0)'
        ws[f"C{r}"] = f'=IF(A{r}=ROUND(Inputs!$B$18,0),MAX($B$11:$B$42)*1.05,NA())'
        ws[f"D{r}"] = f'=IF(A{r}=ROUND(Inputs!$B$14,0),MAX($B$11:$B$42)*1.05,NA())'
        ws[f"B{r}"].number_format = "0.0%"

    for r in range(11, 63):
        idx = r - 11
        ws[f"F{r}"] = f'=ROUND(Inputs!$B$15,0)-25+{idx}'
        ws[f"G{r}"] = f'=IF(AND(ISNUMBER(F{r}),ISNUMBER(Inputs!$B$15),ISNUMBER(Inputs!$B$17),Inputs!$B$17>0),NORMDIST(F{r}+0.5,Inputs!$B$15,Inputs!$B$17,TRUE)-NORMDIST(F{r}-0.5,Inputs!$B$15,Inputs!$B$17,TRUE),0)'
        ws[f"H{r}"] = f'=IF(F{r}=ROUND(Inputs!$B$19,0),MAX($G$11:$G$62)*1.05,NA())'
        ws[f"I{r}"] = f'=IF(F{r}=ROUND(Inputs!$B$15,0),MAX($G$11:$G$62)*1.05,NA())'
        ws[f"G{r}"].number_format = "0.0%"

    for r in range(11, 26):
        idx = r - 11
        ws[f"K{r}"] = f'=ROUND(Inputs!$B$18,0)-7+{idx}'
        ws[f"L{r}"] = f'=IF(AND(ISNUMBER(K{r}),ISNUMBER(Inputs!$B$14),ISNUMBER(Inputs!$B$16),Inputs!$B$16>0),1-NORMDIST(-K{r},Inputs!$B$14,Inputs!$B$16,TRUE),"")'
        ws[f"M{r}"] = f'=IF(AND(ISNUMBER(L{r}),L{r}>0,L{r}<1),1/L{r},"")'
        ws[f"N{r}"] = f'=IF(AND(ISNUMBER(L{r}),L{r}>0,L{r}<1),IF(L{r}>=0.5,-100*L{r}/(1-L{r}),100*(1-L{r})/L{r}),"")'
        ws[f"L{r}"].number_format = "0.0%"
        ws[f"M{r}"].number_format = "0.00"
        ws[f"N{r}"].number_format = "0"

    for r in range(11, 26):
        idx = r - 11
        ws[f"P{r}"] = f'=ROUND(Inputs!$B$19,0)-7+{idx}'
        ws[f"Q{r}"] = f'=IF(AND(ISNUMBER(P{r}),ISNUMBER(Inputs!$B$15),ISNUMBER(Inputs!$B$17),Inputs!$B$17>0),1-NORMDIST(P{r},Inputs!$B$15,Inputs!$B$17,TRUE),"")'
        ws[f"R{r}"] = f'=IF(AND(ISNUMBER(Q{r}),Q{r}>0,Q{r}<1),1/Q{r},"")'
        ws[f"S{r}"] = f'=IF(AND(ISNUMBER(Q{r}),Q{r}>0,Q{r}<1),IF(Q{r}>=0.5,-100*Q{r}/(1-Q{r}),100*(1-Q{r})/Q{r}),"")'
        ws[f"Q{r}"].number_format = "0.0%"
        ws[f"R{r}"].number_format = "0.00"
        ws[f"S{r}"].number_format = "0"

    for row in ws.iter_rows(min_row=10, max_row=62, min_col=1, max_col=19):
        for cell in row:
            cell.border = BORDER

    line1 = LineChart()
    line1.title = "Home Margin Distribution | Dist vs Market vs Model"
    line1.y_axis.title = "Probability Mass"
    line1.x_axis.title = "Home Margin (points)"
    line1.legend.position = "r"
    line1.height = 7
    line1.width = 11
    line1.smooth = True
    line1.add_data(Reference(ws, min_col=2, max_col=4, min_row=10, max_row=42), titles_from_data=True)
    line1.set_categories(Reference(ws, min_col=1, max_col=1, min_row=11, max_row=42))
    ws.add_chart(line1, "J12")

    line2 = LineChart()
    line2.title = "Game Total Distribution | Dist vs Market vs Model"
    line2.y_axis.title = "Probability Mass"
    line2.x_axis.title = "Game Total (points)"
    line2.legend.position = "r"
    line2.height = 7
    line2.width = 11
    line2.smooth = True
    line2.add_data(Reference(ws, min_col=7, max_col=9, min_row=10, max_row=62), titles_from_data=True)
    line2.set_categories(Reference(ws, min_col=6, max_col=6, min_row=11, max_row=62))
    ws.add_chart(line2, "J30")

    ws.freeze_panes = "A10"


def build_dashboard(ws, n_games):
    for col, width in {
        "A": 18, "B": 18, "D": 18, "E": 18, "G": 18, "H": 18,
        "J": 16, "K": 16, "M": 16
    }.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells("A1:K1")
    ws["A1"] = "Market Maker Predictive Pricing Dashboard"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    items = [
        ("A2", "Selected Game", "B2", '=Inputs!$B$5&" @ "&Inputs!$B$4', "@"),
        ("D2", "Selector Row", "E2", '=GamePick_ModeA!$B$2', "0"),

        ("A4", "Market Spread", "B4", '=Inputs!$B$18', "0.0"),
        ("A5", "Model Spread", "B5", '=Inputs!$B$14', "0.0"),
        ("A6", "Spread Edge", "B6", '=Inputs!$B$23', "0.0"),

        ("D4", "Market Total", "E4", '=Inputs!$B$19', "0.0"),
        ("D5", "Model Total", "E5", '=Inputs!$B$15', "0.0"),
        ("D6", "Total Edge", "E6", '=Inputs!$B$24', "0.0"),

        ("G4", "Home Win %", "H4", '=Inputs!$B$20', "0.0%"),
        ("G5", "Home Cover %", "H5", '=Inputs!$B$21', "0.0%"),
        ("G6", "Over %", "H6", '=Inputs!$B$22', "0.0%"),

        ("J4", "Fair ML US", "K4", '=Inputs!$B$26', "0"),
        ("J5", "Fair Cover US", "K5", '=Inputs!$B$28', "0"),
        ("J6", "Fair Over US", "K6", '=Inputs!$B$30', "0"),
    ]
    for lc, ll, vc, fo, nf in items:
        ws[lc] = ll
        ws[lc].fill = TITLE_FILL if lc in ("A2", "D2") else SECTION_FILL
        ws[lc].font = Font(color="FFFFFF", bold=True) if lc in ("A2", "D2") else Font(bold=True)
        ws[lc].border = BORDER
        ws[vc] = fo
        ws[vc].fill = KPI_FILL
        ws[vc].font = Font(bold=True)
        ws[vc].number_format = nf
        ws[vc].border = BORDER

    ws.merge_cells("A9:K9")
    ws["A9"] = "Slate Snapshot"
    ws["A9"].fill = TITLE_FILL
    ws["A9"].font = Font(color="FFFFFF", bold=True)

    snapshot_headers = ["Row", "Away", "Home", "Mkt Spr", "Model Spr", "Edge", "Mkt Tot", "Model Tot", "Edge", "Cover%", "Over%"]
    for i, h in enumerate(snapshot_headers, 1):
        ws.cell(10, i, h)
        ws.cell(10, i).fill = SECTION_FILL
        ws.cell(10, i).font = Font(bold=True)
        ws.cell(10, i).border = BORDER

    max_show = min(n_games, 10)
    for r in range(11, 11 + max_show):
        src = r - 9
        mapping = {
            "A": f"=MarketMaker_Board!A{src}",
            "B": f"=MarketMaker_Board!C{src}",
            "C": f"=MarketMaker_Board!D{src}",
            "D": f"=MarketMaker_Board!F{src}",
            "E": f"=MarketMaker_Board!G{src}",
            "F": f"=MarketMaker_Board!H{src}",
            "G": f"=MarketMaker_Board!I{src}",
            "H": f"=MarketMaker_Board!J{src}",
            "I": f"=MarketMaker_Board!K{src}",
            "J": f"=MarketMaker_Board!M{src}",
            "K": f"=MarketMaker_Board!N{src}",
        }
        for col, formula in mapping.items():
            ws[f"{col}{r}"] = formula
            ws[f"{col}{r}"].border = BORDER
        for col in "DEFGHI":
            ws[f"{col}{r}"].number_format = "0.0"
        ws[f"J{r}"].number_format = "0.0%"
        ws[f"K{r}"].number_format = "0.0%"

    ws["M2"] = "Chart guide"
    ws["M2"].fill = TITLE_FILL
    ws["M2"].font = Font(color="FFFFFF", bold=True)
    ws["M3"] = "Distribution = model probability mass"
    ws["M4"] = "Market Marker = current market number"
    ws["M5"] = "Model Marker = model-implied number"
    for c in ["M3", "M4", "M5"]:
        ws[c].fill = NOTE_FILL
        ws[c].border = BORDER

    line1 = LineChart()
    line1.title = "Home Margin | Distribution vs Market vs Model"
    line1.y_axis.title = "Probability Mass"
    line1.x_axis.title = "Home Margin (points)"
    line1.legend.position = "r"
    line1.height = 7
    line1.width = 11
    line1.smooth = True
    line1.add_data(Reference(ws.parent["SpreadTotal"], min_col=2, max_col=4, min_row=10, max_row=42), titles_from_data=True)
    line1.set_categories(Reference(ws.parent["SpreadTotal"], min_col=1, max_col=1, min_row=11, max_row=42))
    ws.add_chart(line1, "M8")

    line2 = LineChart()
    line2.title = "Game Total | Distribution vs Market vs Model"
    line2.y_axis.title = "Probability Mass"
    line2.x_axis.title = "Game Total (points)"
    line2.legend.position = "r"
    line2.height = 7
    line2.width = 11
    line2.smooth = True
    line2.add_data(Reference(ws.parent["SpreadTotal"], min_col=7, max_col=9, min_row=10, max_row=62), titles_from_data=True)
    line2.set_categories(Reference(ws.parent["SpreadTotal"], min_col=6, max_col=6, min_row=11, max_row=62))
    ws.add_chart(line2, "M24")

    ws.freeze_panes = "A4"


def build_workbook(out_path: Path, game_inputs: pd.DataFrame, blended: pd.DataFrame, team_player: pd.DataFrame, kenpom: pd.DataFrame | None):
    wb = Workbook()
    wb.remove(wb.active)

    dash = add_sheet(wb, "MarketMaker_Dashboard")
    selector = add_sheet(wb, "GamePick_ModeA")
    inputs = add_sheet(wb, "Inputs")
    board = add_sheet(wb, "MarketMaker_Board")
    spread = add_sheet(wb, "SpreadTotal")

    gws = add_sheet(wb, "GameInputs")
    write_df(gws, game_inputs)

    bws = add_sheet(wb, "BlendedRatings")
    write_df(bws, blended)

    tws = add_sheet(wb, "TeamPlayerIndex")
    write_df(tws, team_player)

    if kenpom is not None:
        kws = add_sheet(wb, "KenPomRaw")
        write_df(kws, kenpom)

    n_games = len(game_inputs)
    build_selector(selector, n_games)
    build_inputs(inputs)
    build_board(board, n_games)
    build_spreadtotal(spread)
    build_dashboard(dash, n_games)

    order = [
        "MarketMaker_Dashboard",
        "GamePick_ModeA",
        "Inputs",
        "MarketMaker_Board",
        "SpreadTotal",
        "GameInputs",
        "BlendedRatings",
        "TeamPlayerIndex",
    ]
    if kenpom is not None:
        order.append("KenPomRaw")
    wb._sheets = [wb[s] for s in order]

    for s in ["GameInputs", "BlendedRatings", "TeamPlayerIndex"] + (["KenPomRaw"] if kenpom is not None else []):
        wb[s].sheet_state = "hidden"

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(out_path)
    print(f"Wrote {out_path}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", required=True)
    ap.add_argument("--cache_dir", default="cbb_cache")
    ap.add_argument("--season", type=int, default=2026)
    args = ap.parse_args()

    cache = Path(args.cache_dir)
    gi = pd.read_csv(cache / "GameInputs.csv")
    br = pd.read_csv(cache / "BlendedRatings.csv")
    tp = pd.read_csv(cache / "TeamPlayerIndex.csv")
    kp_path = cache / f"KenPom_Ratings_{args.season}.csv"
    kp = pd.read_csv(kp_path) if kp_path.exists() and kp_path.stat().st_size else None

    build_workbook(Path(args.out), gi, br, tp, kp)


if __name__ == "__main__":
    main()
