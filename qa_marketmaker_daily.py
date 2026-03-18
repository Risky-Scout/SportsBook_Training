#!/usr/bin/env python3
from __future__ import annotations
import argparse, math
from pathlib import Path
import pandas as pd
import openpyxl

def main():
    ap = argparse.ArgumentParser(description="QA check for market-maker daily workbook and cache files.")
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--gameinputs", required=True)
    ap.add_argument("--blended", required=True)
    args = ap.parse_args()

    issues = []

    gi = pd.read_csv(args.gameinputs)
    br = pd.read_csv(args.blended)

    # half-point tick checks
    for col in ["Home spread line (input)", "Game total line (input)"]:
        if col in gi.columns:
            bad = gi[pd.to_numeric(gi[col], errors="coerce").notna() & ((pd.to_numeric(gi[col], errors="coerce")*2).round() != pd.to_numeric(gi[col], errors="coerce")*2)]
            if not bad.empty:
                issues.append(f"{col}: {len(bad)} rows not on half-point ticks")

    # team-name matching
    if "TEAM" in br.columns:
        valid = set(br["TEAM"].astype(str).str.strip())
        for col in ["Home Team","Away Team"]:
            if col in gi.columns:
                bad = sorted(set(gi.loc[~gi[col].astype(str).isin(valid), col].astype(str)))
                if bad:
                    issues.append(f"{col}: unmatched teams -> {bad[:10]}")

    wb = openpyxl.load_workbook(args.workbook, data_only=False, keep_vba=args.workbook.lower().endswith(".xlsm"))
    need = ["Inputs","MarketMaker_Board","SpreadTotal","GameInputs"]
    for s in need:
        if s not in wb.sheetnames:
            issues.append(f"Workbook missing sheet: {s}")

    if all(s in wb.sheetnames for s in ["Inputs","SpreadTotal","GameInputs"]):
        ws = wb["Inputs"]
        for cell in ["B7","B8","B9","B12","B13","B14","B15","B20","B21","B22"]:
            v = ws[cell].value
            if v in (None, ""):
                issues.append(f"Inputs!{cell} blank formula/value")

        st = wb["SpreadTotal"]
        for cell in ["C11","D11","H11","I11"]:
            f = st[cell].value
            if not isinstance(f, str) or "NORMDIST" not in f.upper():
                issues.append(f"SpreadTotal!{cell} not full-curve NORMDIST formula")

        giws = wb["GameInputs"]
        # crude tick check from workbook
        for r in range(2, min(giws.max_row, 30)+1):
            for c in [5,6]:
                v = giws.cell(r,c).value
                if isinstance(v,(int,float)) and abs(v*2-round(v*2))>1e-9:
                    issues.append(f"Workbook GameInputs row {r} col {c} not on half tick: {v}")
                    break

    if issues:
        print("QA FAILED")
        for x in issues:
            print(" -", x)
    else:
        print("QA PASSED")

if __name__ == "__main__":
    main()
